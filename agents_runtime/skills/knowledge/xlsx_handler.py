"""XLSX knowledge handler."""
from __future__ import annotations

import base64
import io
import logging
from typing import Any, Dict, Optional

from skills.knowledge import register_skill

logger = logging.getLogger(__name__)

MIME_TYPES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/excel",
    "application/vnd.ms-excel",
]


async def extract(envelope: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    extra = envelope.get("extra", {})
    mimetype = (
        extra.get("doc_mimetype")
        or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ).lower()
    source_name = extra.get("doc_file_name") or "spreadsheet.xlsx"
    raw_bytes = await _download_bytes(envelope)
    if raw_bytes is None:
        return None
    text = ""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        chunks = []
        for sheet in wb.worksheets:
            chunks.append(f"--- {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                line = " | ".join("" if v is None else str(v) for v in row)
                chunks.append(line)
        text = "\n".join(chunks)
    except Exception as exc:
        logger.warning("xlsx_handler extract failed: %s", exc)
        return None
    return {
        "text": text,
        "source_name": source_name,
        "mimetype": mimetype,
        "raw_size": len(raw_bytes),
    }


async def _download_bytes(envelope: Dict[str, Any]) -> Optional[bytes]:
    extra = envelope.get("extra", {})
    doc_b64 = extra.get("doc_base64", "")
    if doc_b64:
        try:
            return base64.b64decode(doc_b64)
        except Exception:
            pass
    message_id = envelope.get("message_id", "")
    instance = envelope.get("instance", "")
    remote_jid = extra.get("remote_jid", "")
    if message_id and instance:
        try:
            from core.evolution_client import get_base64_from_media_message

            result = await get_base64_from_media_message(
                instance=instance,
                message_id=message_id,
                remote_jid=remote_jid,
            )
            media_b64 = result.get("base64", "")
            if media_b64:
                return base64.b64decode(media_b64)
        except Exception as exc:
            logger.warning("xlsx_handler get_base64 failed: %s", exc)
    return None


async def persist(
    envelope: Dict[str, Any],
    extracted: Dict[str, Any],
    scope: str = "private",
    metadata=None,
) -> Dict[str, Any]:
    phone = envelope.get("phone", "")
    metadata = metadata or {}
    text = extracted.get("text", "")
    source_name = extracted.get("source_name", "spreadsheet.xlsx")
    mimetype = extracted.get("mimetype", MIME_TYPES[0])
    extra_metadata = {"mimetype": mimetype, "scope": "private", **metadata}
    if not text:
        return {"error": "no_text_extracted", "mimetype": mimetype}

    try:
        from core.rag import index_private_document

        result = await index_private_document(
            phone=phone,
            text_content=text,
            source_title=source_name,
            category=extra_metadata.get("class", "legislacao") if extra_metadata.get("class") else "whatsapp_attachment",
            metadata=extra_metadata,
            class_=extra_metadata.get("class"),
            group=extra_metadata.get("group"),
            theme=extra_metadata.get("theme"),
        )
        if result.get("error"):
            return {"error": "rag_index_failed", "detail": result.get("error")}
        if result.get("partial"):
            chunks_idx = result.get("chunks_indexed", 0)
            total = result.get("chunks", 0)
            return {
                "status": "rag_individual_partial",
                "index_result": result,
                "source_name": source_name,
                "scope": "private",
                "category": metadata or {},
                "chunks_indexed": chunks_idx,
                "chunks_total": total,
            }
        chunks_indexed = result.get("chunks_indexed", result.get("chunks", 0))
        return {
            "status": "rag_individual",
            "index_result": result,
            "source_name": source_name,
            "scope": "private",
            "category": metadata,
            "chunks_indexed": chunks_indexed,
        }
    except Exception as exc:
        logger.warning("xlsx_handler private index failed: %s", exc)
        return {"error": "rag_index_failed", "detail": str(exc)}


register_skill("xlsx", type("XLSXSkill", (), {
    "MIME_TYPES": MIME_TYPES,
    "extract": staticmethod(extract),
    "persist": staticmethod(persist),
})())

__all__ = ["extract", "persist", "MIME_TYPES"]
