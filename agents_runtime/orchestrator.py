"""Orchestrator - routes incoming messages to appropriate agent.

Flow:
1. Receive message from /chat endpoint
2. Detect special cases (audio, groups, morality, learning)
3. Determine orchestrator agent (jennifier)
4. Check if message has keywords/intent for direct delegation
5. Call LLM with orchestrator system_prompt + tools
6. If LLM emits function_call, delegate to manager/specialist
7. Return final response with delay_ms

This is a simplified orchestrator (not full Agno Team) for clarity and testability.
"""
import os
import re
import json
import copy
import base64
import logging
import time
import asyncio
import unicodedata
import io
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, Optional, List

from core.llm_provider import LLMProvider, LLMError
from core.masker import mask_pii
from core.delay_calculator import calculate_delay_ms, calculate_presence
from core.commands import detect_command, apply_command
from core.tabular import (
    build_calendar_payload,
    build_drive_payload,
    build_email_payload,
)
from tool_registry import get_tool, get_tool_schema, is_user_scoped_tool
from agent_loader import get_agent, get_skill, has_nickname
from core.audit import log_action
from core.timezone import now_brt
from tools.api_registry import api_registry
from core.observability import (
    new_tracker, set_current_tracker, current_tracker,
)

logger = logging.getLogger(__name__)

_interaction_history: List[Dict[str, Any]] = []
MAX_HISTORY = 20
_response_cache: Dict[str, Dict[str, Any]] = {}
CACHE_TTL_SEC = int(os.getenv("RESPONSE_IDEMPOTENCY_TTL_SEC", "86400"))
_indexing_tasks: set = set()

MULTI_SPECIALIST_TIMEOUT_SEC = float(os.getenv("MULTI_SPECIALIST_TIMEOUT_SEC", "30"))
PREFETCH_DRIVE_MULTI_TIMEOUT_SEC = float(os.getenv("PREFETCH_DRIVE_MULTI_TIMEOUT_SEC", "8"))

LLM_MAX_TOKENS_MANAGER = int(os.getenv("LLM_MAX_TOKENS_MANAGER", "1500"))
LLM_MAX_TOKENS_DEFAULT = int(os.getenv("LLM_MAX_TOKENS_DEFAULT", "500"))


def _finish_indexing_task(task: asyncio.Task) -> None:
    _indexing_tasks.discard(task)
    if task.cancelled():
        logger.warning("RAG indexing task cancelled")
        return
    exception = task.exception()
    if exception:
        logger.error("RAG indexing task failed: %s", exception)


def _schedule_indexing(coroutine: Any) -> asyncio.Task:
    task = asyncio.create_task(coroutine)
    _indexing_tasks.add(task)
    task.add_done_callback(_finish_indexing_task)
    return task


async def drain_indexing_tasks(timeout: float = 15.0) -> None:
    pending = list(_indexing_tasks)
    if not pending:
        return
    _, unfinished = await asyncio.wait(pending, timeout=timeout)
    if unfinished:
        logger.warning("RAG indexing drain timed out: pending=%s", len(unfinished))


def _message_id(payload: Dict[str, Any]) -> Optional[str]:
    extra = payload.get("extra", {})
    message_id = (
        payload.get("message_id")
        or extra.get("message_id")
        or extra.get("messageId")
        or (extra.get("key") or {}).get("id")
    )
    if not message_id:
        phone = re.sub(r"\D", "", str(payload.get("phone", "")))
        owner_hash = __import__("hashlib").sha256(phone.encode("utf-8")).hexdigest()[:12] if phone else "unknown"
        logger.warning(
            "message_id_missing owner_hash=%s instance=%s fallback=time_based_key retry_idempotent=false",
            owner_hash,
            payload.get("instance", "unknown"),
        )
    return message_id or None


def _conversation_id(payload: Dict[str, Any]) -> str:
    extra = payload.get("extra", {})
    return str(extra.get("remote_jid") or extra.get("conversation_id") or payload.get("phone", ""))


def _idempotency_key(payload: Dict[str, Any]) -> Optional[str]:
    message_id = _message_id(payload)
    if not message_id:
        return None
    raw = f"{payload.get('instance', 'jennifer')}:{_conversation_id(payload)}:{message_id}"
    return __import__("hashlib").sha256(raw.encode("utf-8")).hexdigest()


async def _finalize_orchestration(
    payload: Dict[str, Any],
    masked_text: str,
    sender_name: str,
    result: Dict[str, Any],
    path: List[Dict[str, Any]],
    cache_key: Optional[str],
) -> Dict[str, Any]:
    metadata = result.setdefault("metadata", {})
    metadata.setdefault("response_identity", "Jennifer")
    try:
        from core.observability import attach_to_metadata, current_tracker
        attach_to_metadata(metadata, current_tracker())
    except Exception as exc:
        logger.debug("observability_attach_failed: %s", exc)
    path.append({
        "step": 3,
        "phase": "result",
        "agent_id": metadata.get("agent_id"),
        "model": metadata.get("model_used"),
        "escalated": metadata.get("escalated"),
        "confidence": metadata.get("confidence_score"),
        "tool_rounds": metadata.get("tool_rounds", 0),
        "tool_calls": metadata.get("tool_calls", []),
        "response_identity": metadata.get("response_identity"),
    })
    phone = payload.get("phone", "")
    _interaction_history.append({
        "timestamp": now_brt().isoformat(),
        "phone": phone,
        "text_preview": masked_text[:80],
        "sender": mask_pii(sender_name),
        "path": path,
        "reply_preview": mask_pii(result.get("reply", ""))[:80],
    })
    if len(_interaction_history) > MAX_HISTORY:
        _interaction_history.pop(0)

    if cache_key and not metadata.get("error"):
        cached_result = copy.deepcopy(result)
        cached_result["ts"] = int(time.time())
        _response_cache[cache_key] = cached_result
        for key in list(_response_cache.keys()):
            if int(time.time()) - _response_cache.get(key, {}).get("ts", 0) > CACHE_TTL_SEC:
                del _response_cache[key]

    message_id = _message_id(payload)
    conversation_id = _conversation_id(payload)
    agent_id = metadata.get("agent_id", "jennifier")
    _schedule_indexing(_index_message(
        phone,
        masked_text,
        "in",
        message_id=message_id,
        conversation_id=conversation_id,
        turn_id=message_id,
        agent_id="user",
        response_identity="Usuario",
    ))
    reply_text = result.get("reply", "")
    if reply_text:
        _schedule_indexing(_index_message(
            phone,
            reply_text,
            "out",
            message_id=f"{message_id}:reply" if message_id else None,
            conversation_id=conversation_id,
            turn_id=message_id,
            agent_id=agent_id,
            response_identity="Jennifer",
        ))

    if (
        reply_text
        and payload.get("phone")
    ):
        try:
            tabular_payload = _detect_tabular_payload(result)
            user_force_image = _user_requested_image(masked_text)
            auto_image_enabled = os.getenv("IMAGE_REPORT_AUTO", "false").lower() == "true"
            skip_image = bool(metadata.get("skip_image_report")) and not user_force_image
            # Guardrail Anti-Duplicação: Envia imagem apenas se explicitamente requisitada pelo usuário
            # ou se IMAGE_REPORT_AUTO=true configurado no ambiente.
            if tabular_payload and not skip_image and (user_force_image or auto_image_enabled):
                sent_ok = await _auto_send_image(payload, tabular_payload, reply_text)
                if sent_ok:
                    # Guardrail: Se a imagem com legenda já foi despachada, suprime o envio de texto duplicado
                    result["delivered_as_image"] = True
                    result["original_reply"] = reply_text
                    result["reply"] = ""
        except Exception as exc:
            logger.warning("auto_render_failed: %s", exc)

    try:
        result = await _maybe_onboarding_nudge(payload, result)
    except Exception as exc:
        logger.debug("onboarding_nudge_skipped: %s", exc)

    return result


async def _user_groups_context(phone: str) -> str:
    """Gera contexto de grupos em comum (G3): 'você e X estão nos grupos Y, Z'.

    Le o indice denormalizado ``usuarios/{phone}.group_memberships`` (1
    doc.get) com cache in-memory TTL (core.user_groups_cache). Retorna ""
    quando nao ha grupos ou erro. Substitui a collection-group query legada
    (group_members/member_phones array_contains), que nao tinha indice.
    """
    try:
        from core.message_ledger import _get_firestore
        from core import user_groups_cache
        from agent_loader import _canonical_phone

        digits = "".join(c for c in str(phone or "") if c.isdigit())
        if not digits:
            return ""
        canonical = _canonical_phone(digits) or digits

        cached = user_groups_cache.get(canonical)
        if cached is not None:
            logger.debug("user_groups_context_cache_hit phone=%s", canonical)
            return cached

        db = _get_firestore()
        if db is None:
            return ""

        def fetch() -> list:
            doc = db.collection("usuarios").document(canonical).get()
            if not doc.exists:
                return []
            return (doc.to_dict() or {}).get("group_memberships") or []

        memberships = await asyncio.to_thread(fetch)

        groups = []
        for m in memberships or []:
            if isinstance(m, dict):
                groups.append(m.get("subject") or m.get("gid") or "")

        ctx = (
            f"Grupos em comum com o usuario: {', '.join(g for g in groups if g)}"
            if groups
            else ""
        )
        user_groups_cache.set(canonical, ctx)
        return ctx
    except Exception as exc:  # noqa: BLE001
        logger.debug("user_groups_context_skipped phone=%s exc=%s", phone, exc)
        return ""


def _get_group_name_by_jid(phone: str, group_jid: str) -> str:
    """Busca o nome (subject) de um grupo a partir de group_jid e phone."""
    if not group_jid or "@g.us" not in group_jid:
        return ""
    try:
        from agent_loader import get_user, _canonical_phone
        canonical = _canonical_phone(phone) or phone
        user = get_user(canonical) or {}
        memberships = user.get("group_memberships") or []
        for m in memberships:
            if isinstance(m, dict) and m.get("gid") == group_jid:
                return m.get("subject") or ""
        from core.message_ledger import _get_firestore
        db = _get_firestore()
        if db is not None:
            doc = db.collection("group_members").document(group_jid.replace("/", "_")).get()
            if doc.exists:
                return doc.to_dict().get("subject") or ""
    except Exception as exc:  # noqa: BLE001
        logger.debug("get_group_name_by_jid error: %s", exc)
    return ""


async def _resolve_group_mentions(payload: Dict[str, Any]) -> str:
    """Resolve @LID mencionados numa mensagem de grupo para os NOMES.

    Quando alguem digita '@Clarissa Pontual', o WhatsApp grava o LID
    (ex: @210870093217996) no texto e em contextInfo.mentionedJid. Esta
    funcao consulta o snapshot group_members e devolve os nomes reais,
    para a Jennifer saber quem foi mencionado sem expor telefone.
    """
    try:
        extra = payload.get("extra", {}) or {}
        remote_jid = str(payload.get("remote_jid") or extra.get("remote_jid") or "")
        if "@g.us" not in remote_jid:
            return ""
        mentioned = extra.get("mentioned_jids") or []
        if not mentioned:
            return ""
        from tools.group import resolve_mentioned

        resolved = resolve_mentioned(remote_jid, mentioned)
        if not resolved:
            return ""

        instance = payload.get("instance") or extra.get("instance") or "Jennifer"
        from core.evolution_webhook import _resolve_bot_lid, _resolve_bot_jid
        bot_lid = _resolve_bot_lid(instance, remote_jid) or ""
        bot_jid = _resolve_bot_jid(instance) or ""
        bot_lid_raw = bot_lid.split("@")[0] if bot_lid else ""
        bot_jid_raw = bot_jid.split("@")[0] if bot_jid else ""

        names = []
        for m in resolved:
            if not m:
                continue
            lid_raw = str(m.get("lid") or "").split("@")[0]
            phone_raw = str(m.get("phone") or "").split("@")[0]
            if (bot_lid_raw and lid_raw == bot_lid_raw) or (bot_jid_raw and phone_raw == bot_jid_raw):
                continue  # Pula o proprio bot (Jennifer)
            names.append(m["name"] or m["phone"])

        if not names:
            return ""
        return "Pessoas mencionadas nesta mensagem (alem de voce): " + ", ".join(names)
    except Exception as exc:  # noqa: BLE001
        logger.debug("resolve_group_mentions_skipped exc=%s", exc)
        return ""


async def _user_has_any_connection(phone: str):
    """True se o user ja tem OAuth Google OU alguma conexao Composio.

    Retorna None quando nao e possivel verificar (Firestore/Composio
    indisponiveis) — nesse caso o onboarding nudge NAO deve disparar,
    para nao anexar dica em respostas de usuarios ja conectados.
    """
    try:
        from agent_loader import get_user

        user = get_user(phone) or {}
        if user.get("google_oauth_token"):
            return True
        from tools.composio_connect import get_status

        status = await get_status(phone)
        apps = (status or {}).get("apps") or {}
        if any((a or {}).get("connected") for a in apps.values()):
            return True
        return False
    except Exception:  # noqa: BLE001
        return None


_LAST_NUDGE_TS: Dict[str, float] = {}


async def _maybe_onboarding_nudge(payload: Dict[str, Any], result: Dict[str, Any]) -> Dict[str, Any]:
    """Onboarding (P5/loop 12/08/2026): user novo sem conexao recebe link limpo.

    Apenas conversa PRIVADA (nao grupo), sem erro na resposta, sem link duplicado,
    e respeitando debounce de 60s.
    """
    extra = payload.get("extra", {}) or {}
    if extra.get("is_group"):
        return result
    phone = str(payload.get("phone", "") or "")
    if not phone:
        return result

    # Debounce de 60s para evitar flood de links na mesma sessao
    now = time.time()
    last_ts = _LAST_NUDGE_TS.get(phone, 0.0)
    if now - last_ts < 60.0:
        return result

    metadata = result.get("metadata", {}) or {}
    reply_lower = result.get("reply", "").lower()
    if (
        metadata.get("error")
        or metadata.get("blocked")
        or metadata.get("blocked_reason")
        or "conecte suas contas" in reply_lower
        or "oauth/google" in reply_lower
        or "/portal" in reply_lower
        or "tinyurl.com" in reply_lower
    ):
        return result
    has_conn = await _user_has_any_connection(phone)
    if has_conn is not False:
        return result  # conectado OU nao verificavel (None) — sem nudge
    reply = result.get("reply", "")
    if not reply:
        return result
    url = _onboarding_url(phone)
    _LAST_NUDGE_TS[phone] = now
    nudge = (
        f"\n\n💡 *Dica:* para eu acessar seus emails, agenda, Drive, "
        f"YouTube, LinkedIn e mais, conecte suas contas aqui: {url}"
    )
    result["reply"] = reply + nudge
    return result


def _detect_tabular_payload(result: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Inspect agent metadata for tabular payloads we can render as a PNG.

    Returns a dict ready to feed ``render_report`` or None.

    Ordem de detecção:
    1. ``metadata["tabular"]`` (anexado pelo ``pipelines/_executor.run_agent``
       quando o prefetch retornou dados estruturados - calendar/email/drive).
    2. ``metadata["tool_results"]`` (quando o LLM chama uma tool diretamente
       e o resultado é tabulável: drive.list_folder, gmail.search_messages,
       calendar.list_events, knowledge.retrieve).
    """
    metadata = result.get("metadata", {}) or {}

    # 1) Tabular anexado pelo prefetch (pipeline com `tools: []` - dados injetados).
    prefetch_tabular = metadata.get("tabular")
    if isinstance(prefetch_tabular, dict) and prefetch_tabular.get("rows"):
        return prefetch_tabular

    # 2) Tool results diretos do LLM.
    tool_results = metadata.get("tool_results") or []
    if not isinstance(tool_results, list):
        return None

    for entry in tool_results:
        if not isinstance(entry, dict):
            continue
        tool = entry.get("tool") or entry.get("name") or ""
        result_data = entry.get("result")
        if not isinstance(result_data, dict):
            continue
        if tool == "drive.list_folder" or tool.endswith("list_folder"):
            payload = build_drive_payload(result_data.get("files") or [])
            if payload:
                return payload
        if tool == "gmail.search_messages" or tool.endswith("search_messages"):
            messages = result_data.get("messages") or result_data.get("threads") or []
            payload = build_email_payload(messages)
            if payload:
                return payload
        if tool == "calendar.list_events" or tool.endswith("list_events"):
            events = result_data.get("events") or result_data.get("items") or []
            payload = build_calendar_payload(events)
            if payload:
                return payload
        if tool == "knowledge.retrieve" or tool.endswith("knowledge.retrieve") or tool.endswith("retrieve_knowledge"):
            chunks = result_data.get("results") or []
            count = result_data.get("count", len(chunks))
            if not chunks or count == 0:
                continue
            rows = []
            for c in chunks[:5]:
                source = str(
                    c.get("source_title") or c.get("source") or c.get("metadata_filename") or "-"
                )[:48]
                excerpt = str(
                    c.get("content") or c.get("text") or c.get("snippet") or ""
                )[:120].replace("\n", " ")
                score = f"{c.get('score', 0.0):.2f}" if c.get("score") is not None else "-"
                rows.append([source, excerpt, score])
            return {
                "title": f"Conhecimento encontrado ({count} trechos)",
                "headers": ["Fonte", "Trecho", "Score"],
                "rows": rows,
                "emoji_header": "📚",
            }
    return None


_IMAGE_REQUEST_KEYWORDS = ("tabela", "imagem", "grafico", "gráfico", "png", "print", "planilha")


def _user_requested_image(masked_text: str) -> bool:
    """Detecta pedido explicito do usuario por render visual.

    Usado pelo auto-image para sobrescrever ``skip_image_report`` quando
    o usuario pede "em tabela", "como imagem", "grafico" etc. A normalizacao
    usa ``unicodedata.normalize('NFKD', ...)`` para casar "grafico"/"gráfico".
    """
    if not masked_text:
        return False
    normalized = "".join(
        c for c in unicodedata.normalize("NFKD", masked_text.lower())
        if not unicodedata.combining(c)
    )
    return any(kw in normalized for kw in _IMAGE_REQUEST_KEYWORDS)


async def _auto_send_image(
    payload: Dict[str, Any],
    tabular: Dict[str, Any],
    caption: str,
) -> bool:
    """Render and dispatch the tabular payload as a PNG to WhatsApp."""
    from core.evolution_client import send_image
    from tools.image_report import render_report

    rendered = render_report(
        title=tabular["title"],
        headers=tabular.get("headers"),
        rows=tabular.get("rows", []),
        emoji_header=tabular.get("emoji_header", ""),
        footer=caption[:120],
    )
    if not rendered or rendered.get("error"):
        return False
    extra = payload.get("extra", {}) or {}
    instance = payload.get("instance", "Jennifer")
    phone = payload.get("phone", "")
    if not phone:
        return False
    try:
        await send_image(
            instance=instance,
            phone=phone,
            image_bytes=rendered["png_bytes"],
            filename=f"{tabular['title'].lower().replace(' ', '_')}.png",
            caption=caption[:1024],
            remote_jid=extra.get("remote_jid", ""),
        )
        return True
    except Exception as exc:
        logger.warning("auto_send_image_failed: %s", exc)
        return False


def get_recent_interactions(limit: int = 5) -> List[Dict[str, Any]]:
    """Return the most recent orchestration interactions."""
    return _interaction_history[-limit:]

GROSS_KEYWORDS = [
    "puta", "merda", "caralho", "fdp", "porra",
    "buceta", "viado", "bicha", "desgraça",
    "foder", "fode", "piranha", "vagabunda", "puto",
    "bosta", "porcaria", "desgraçado",
]
ASSAULT_KEYWORDS = [
    "assedio", "abuso", "estupro", "violencia", "agressao",
    "ameaça", "ameaca", "chantagem",
]
CORRECTION_KEYWORDS = [
    "na verdade", "não é assim", "nao e assim", "errado", "errada",
]
CALENDAR_KEYWORDS = [
    "agenda", "agend", "reuniao", "evento", "eventos", "compromisso",
    "compromissos", "lembrete", "calendario", "disponivel",
    "semana que vem", "proxima semana", "agenda de hoje",
]
DRIVE_KEYWORDS = [
    "drive", "gdrive", "google drive",
    "meu drive", "no drive", "no gdrive", "meu gdrive",
    "salvar no drive", "salvar no gdrive", "salva no drive",
    "guardar no drive", "guarda no drive", "quero no meu drive",
    "quero no gdrive", "manda pra mim", "envia pra mim",
    "lista os arquivos", "liste os arquivos", "lista de arquivos",
    "mostrar arquivos", "lista os arquivos do drive",
    "dentro desse drive", "nesse drive", "dentro desse gdrive",
    "nesse gdrive", "dentro do drive",
]
DRIVE_KEYWORDS_REMOVED = [
    "documento", "documentos", "arquivo", "arquivos",
    "pasta", "upload", "omnichannel", "atividades", "baixar",
    "encontrar arquivo", "meus arquivos", "meus documentos",
    "buscar arquivo", "buscar arquivos", "procurar documento", "ata", "minuta", "relatorio",
    "apresentação", "apresentacao", "docx", "pdf", "xlsx", "planilha",
    "leia o arquivo", "leia a ata", "abra o arquivo",
    "ache a ata", "procure a ata",
]
EMAIL_KEYWORDS = [
    "email", "e-mail", "emails", "e-mails",
    "caixa de entrada", "caixa postal", "correio", "inbox",
    "gmail", "ler email", "enviar email", "ultimos emails",
    "ultima mensagem", "mensagens",
]
WEB_KEYWORDS = [
    "pesquisar", "buscar na internet", "busque na internet", "procure na web",
    "pesquise na web", "noticia atual", "noticias atuais", "pesquisa sobre",
]
RUNTIME_STATUS_KEYWORDS = [
    "quantos agentes", "quais agentes", "agentes funcionando", "agentes ativos",
    "agentes rodando", "status dos agentes", "o que esta rodando", "o que está rodando",
    "listar agentes", "liste os agentes",
]
AMBIGUOUS_WEB_KEYWORDS = {"o que e", "quem e", "significa"}
INTIMACY_KEYWORDS = [
    "me chame de", "pode me chamar de", "meu apelido",
    "meu nome e", "meu nome é", "como devo te chamar",
]


def _normalize_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(text or ""))
    without_accents = "".join(character for character in normalized if not unicodedata.combining(character))
    return re.sub(r"\s+", " ", without_accents.lower()).strip()


# def _matches_keyword(text: str, keyword: str) -> bool:
#     normalized_keyword = _normalize_text(keyword)
#     if not normalized_keyword:
#         return False
#     pattern = rf"(?<!\w){re.escape(normalized_keyword)}s?(?!\w)"
#     return re.search(pattern, text) is not None


# def _get_routing_rules() -> List[Dict[str, Any]]:
#     config = get_config("routing")
#     rules = []
#     for source in (config or {}).get("rules", []):
#         if not source.get("enabled", True):
#             continue
#         rule = dict(source)
#         if rule.get("agent_id") == "manager-web":
#             rule["keywords"] = [
#                 keyword for keyword in rule.get("keywords", [])
#                 if _normalize_text(keyword) not in AMBIGUOUS_WEB_KEYWORDS
#             ]
#         rules.append(rule)
#     return rules


def _attachment_pending_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    extra = payload.get("extra", {})
    persisted_extra = {
        key: extra.get(key)
        for key in (
            "has_document",
            "doc_mimetype",
            "doc_file_name",
            "doc_file_length",
            "remote_jid",
            "is_group",
        )
        if extra.get(key) is not None
    }
    return {
        "instance": payload.get("instance", "jennifer"),
        "phone": payload.get("phone", ""),
        "message_id": payload.get("message_id", ""),
        "sender_name": payload.get("sender_name", "user"),
        "extra": persisted_extra,
    }


async def _handle_attachment(
    payload: Dict[str, Any],
    intent: Dict[str, Any],
    sender_name: str,
) -> Optional[Dict[str, Any]]:
    """F4d: handler centralizado de attachments (PDF, DOCX, XLSX, etc.).

    Chamado pelo orchestrate() quando extra.has_document=True E intent
    is_attachment=True. Decide entre:
    - ambíguo (sem is_attachment_save nem is_attachment_file): pergunta
    - salvar (is_attachment_file=True): upload_file no Drive
    - memorizar (is_attachment_save=True): index_group_document (grupo)
      ou index_private_document (individual) — F4'

    Returns o result dict pronto para _finalize_orchestration, ou None
    se o handler nao deve ser executado.
    """
    from core.delay_calculator import calculate_delay_ms
    from core.evolution_client import send_text

    extra = payload.get("extra", {})
    phone = payload.get("phone", "")
    instance = payload.get("instance", "")

    is_attachment = bool(intent.get("is_attachment"))
    if not is_attachment:
        return None

    is_save = bool(intent.get("is_attachment_save"))
    is_file = bool(intent.get("is_attachment_file"))
    is_ambiguous = not (is_save or is_file)

    _RAG_KEYWORDS = (
        "memorizar", "memorize", "indexar", "indexe",
        "base de conhecimento", "banco semantico",
        "armazenar na base", "guardar na base",
        "no rag", "no vector", "no firestore",
    )
    _DRIVE_KEYWORDS = (
        "salvar no drive", "guardar no drive",
        "subir no drive", "gdrive", "google drive",
        "meu drive", "no drive", "salva no drive",
        "upload", "faz upload",
    )

    if is_ambiguous:
        caption = (payload.get("text", "") or "").lower()
        if any(kw in caption for kw in _RAG_KEYWORDS):
            is_save = True
            is_ambiguous = False
        elif any(kw in caption for kw in _DRIVE_KEYWORDS):
            is_file = True
            is_ambiguous = False

    async def _send_ack(text: str) -> None:
        try:
            await send_text(
                instance=instance,
                phone=phone,
                text=text,
                delay_ms=0,
                presence="composing",
                remote_jid=extra.get("remote_jid", ""),
            )
        except Exception:
            pass

    if is_ambiguous:
        from core.pending_actions import set_pending_action

        await set_pending_action(
            phone,
            "attachment_mode",
            {"attachment_payload": _attachment_pending_payload(payload)},
            ttl_sec=300,
        )
        await _send_ack(
            "Esse arquivo e para memorizar no banco semantico ou salvar no drive? "
            "Responda 'memorizar' ou 'salvar'."
        )
        reply = "Aguardando confirmacao sobre o arquivo."
        return {
            "reply": reply,
            "delay_ms": calculate_delay_ms(reply),
            "presence": "paused",
            "metadata": {
                "agent_id": "document-handler",
                "response_identity": "Jennifer",
                "waiting_confirmation": "attachment_mode",
            },
        }

    extracted = await _extract_text_from_attachment(payload)
    if not extracted or not extracted.get("text"):
        reply = "Nao consegui extrair texto desse arquivo. Pode tentar de outro formato?"
        return {
            "reply": reply,
            "delay_ms": calculate_delay_ms(reply),
            "presence": "paused",
            "metadata": {
                "agent_id": "document-handler",
                "response_identity": "Jennifer",
                "error": "text_extraction_failed",
            },
        }

    # Consolidado para 1 unico ack antes da indexação pesada
    await _send_ack("ok. pode deixar, estou memorizando o conteudo")

    save_to_rag = is_save
    logger.info(
        "attachment_routing phone=%s decision=%s save_to_rag=%s source=%s",
        phone,
        "rag" if save_to_rag else "drive",
        save_to_rag,
        extracted.get("source_name", "?"),
    )
    persist = await _persist_attachment(payload, extracted, save_to_rag)
    if persist.get("error"):
        reply = (
            f"Tive problema ao salvar: {persist.get('error')}. "
            "Pode tentar de novo?"
        )
        return {
            "reply": reply,
            "delay_ms": calculate_delay_ms(reply),
            "presence": "paused",
            "metadata": {
                "agent_id": "document-handler",
                "response_identity": "Jennifer",
                "error": persist.get("error"),
                "detail": persist.get("detail", ""),
            },
        }

    status = persist.get("status", "")
    source_name = extracted.get("source_name", "document")
    if status in {"rag_group", "rag_individual", "rag_individual_partial"}:
        index_result = persist.get("index_result", {})
        indexed = persist.get("chunks_indexed", index_result.get("chunks_indexed", index_result.get("chunks", index_result.get("indexed", 0))))
        if indexed == 0:
            reply = (
                f"Tive um problema ao memorizar os trechos do arquivo '{source_name}'. "
                "Pode tentar reenviar para que eu tente novamente?"
            )
        elif status == "rag_individual_partial":
            chunks_idx = persist.get("chunks_indexed", indexed)
            chunks_total = persist.get("chunks_total", indexed)
            reply = (
                f"Feito! Indexei {chunks_idx}/{chunks_total} trechos do "
                f"arquivo '{source_name}' na sua base de conhecimento (alguns embeddings falharam, "
                f"mas o doc esta parcialmente pesquisavel). Quer me perguntar algo?"
            )
        else:
            reply = (
                f"Feito! Memorei {indexed} trechos do arquivo '{source_name}' "
                "na sua base de conhecimento. Quer me perguntar algo sobre o arquivo para verificar?"
            )
        try:
            from agent_orchestration.knowledge_retriever import register_indexing
            register_indexing(phone)
        except Exception:
            pass
    elif status.startswith("drive_"):
        folder = "Meu Drive" if status == "drive_individual" else "pasta do grupo"
        reply = (
            f"Feito! 💾 Salvei o arquivo '{source_name}' no {folder}. "
            "Quer me perguntar algo sobre o arquivo para verificar?"
        )
    else:
        reply = "Feito! Arquivo processado."

    return {
        "reply": reply,
        "delay_ms": calculate_delay_ms(reply),
        "presence": "composing",
        "metadata": {
            "agent_id": "document-handler",
            "response_identity": "Jennifer",
            "attachment": status,
            "source_name": source_name,
        },
    }


# def _detect_intent(text: str) -> Dict[str, Any]:
#     normalized = _normalize_text(text)
#     explicit_url = re.search(r"https?://\S+", str(text or ""), flags=re.IGNORECASE) is not None
#     intent = {
#         "is_runtime_status": any(_matches_keyword(normalized, keyword) for keyword in RUNTIME_STATUS_KEYWORDS),
#         "is_gross": any(_matches_keyword(normalized, keyword) for keyword in GROSS_KEYWORDS),
#         "is_assault_related": any(_matches_keyword(normalized, keyword) for keyword in ASSAULT_KEYWORDS),
#         "is_correction": any(_matches_keyword(normalized, keyword) for keyword in CORRECTION_KEYWORDS),
#         "is_calendar": any(_matches_keyword(normalized, keyword) for keyword in CALENDAR_KEYWORDS),
#         "is_drive": any(_matches_keyword(normalized, keyword) for keyword in DRIVE_KEYWORDS),
#         "is_email": any(_matches_keyword(normalized, keyword) for keyword in EMAIL_KEYWORDS),
#         "is_web_search": explicit_url or any(_matches_keyword(normalized, keyword) for keyword in WEB_KEYWORDS),
#         "is_intimacy": any(_matches_keyword(normalized, keyword) for keyword in INTIMACY_KEYWORDS),
#     }
#     attachment_save_kw = ("memorize", "memorizar", "memorizando", "guarde", "guardar",
#                          "indexe", "indexar", "salve na base", "base de conhecimento",
#                          "knowledge base", "no conhecimento", "salva no conhecimento",
#                          "armazene", "armazenar", "cache isso", "no vector", "no firestore",
#                          "leia isso", "leia e armazene", "leia e guarde",
#                          "salve isso", "salvar isso", "guarda isso", "guardar isso",
#                          "indexe isso", "indexar isso", "memorize isso", "memorizar isso",
#                          "conteudo do pdf", "conteudo do docx", "conteudo do xlsx",
#                          "conteudo da pasta", "conteudo do arquivo", "conteudo deste",
#                          "conteudo do doc", "pdf", "docx", "xlsx", "txt",
#                          "anexei", "anexo", "salvar no drive", "guarda no drive")
#     attachment_file_kw = ("so salve", "só salve", "salva o arquivo", "salvar arquivo",
#                           "guarda o arquivo", "manda pra mim", "envia pra mim",
#                           "apenas salve", "nao memorize", "não memorize",
#                           "so guarda", "só guarda", "apenas guarda",
#                           "conteudo do arquivo", "conteudo da pasta",
#                           "anexei", "anexo", "arquivo anexo", "pdf anexo")
#     intent["is_attachment_save"] = any(_matches_keyword(normalized, keyword)
#                                        for keyword in attachment_save_kw)
#     intent["is_attachment_file"] = any(_matches_keyword(normalized, keyword)
#                                        for keyword in attachment_file_kw)
#     intent["is_attachment"] = intent["is_attachment_save"] or intent["is_attachment_file"]
    # F4d.1: se has_document=True e nenhum flag setado, forçar is_attachment=True
    # para que o handler de attachment seja chamado e o user possa
    # responder com 'memorizar' ou 'salvar'.
#     _has_doc = False
    # Checar através de payload, não de kwargs — heurística leve
#     intent["is_attachment"] = intent["is_attachment"] or _has_doc
#     for rule in _get_routing_rules():
#         agent_id = rule.get("agent_id", "")
#         keywords = rule.get("keywords", [])
#         if any(_matches_keyword(normalized, keyword) for keyword in keywords):
#             intent[f"matched_{agent_id}"] = True
#     return intent


def _build_skills_section(skill_ids: List[str]) -> str:
    """Build skills content section for system prompt."""
    if not skill_ids:
        return ""
    parts = ["\n\n# Skills ativas:"]
    for sid in skill_ids:
        skill = get_skill(sid)
        if skill and skill.get("enabled", True):
            parts.append(f"\n## {skill['name']}\n{skill.get('content', '')}")
    return "\n".join(parts)


# async def _get_orchestrator(instance: str) -> Optional[str]:
#     """Get orchestrator with cold-start retry."""
#     orchestrator_id = _select_orchestrator_agent(instance)
#     if not orchestrator_id:
#         await asyncio.sleep(3)
#         orchestrator_id = _select_orchestrator_agent(instance)
#     return orchestrator_id


def _extract_first_name(sender_name: str) -> str:
    """Extrai o primeiro nome do sender_name."""
    if not sender_name or sender_name == "user":
        return ""
    parts = sender_name.strip().split()
    return parts[0] if parts else sender_name


# def _select_orchestrator_agent(instance: str) -> Optional[str]:
#     """Select which orchestrator agent to use for this instance."""
#     for agent_id, agent in _iter_agents():
#         if agent.get("role") == "orchestrator" and agent.get("enabled", True):
#             if instance.lower() in [i.lower() for i in agent.get("instances", [])] or not agent.get("instances"):
#                 return agent_id
#     return None


# def _agent_has_tool(agent_id: str, tool_prefix: str) -> bool:
#     """Return True when the agent's ``tools`` list contains any tool
#     whose id starts with ``tool_prefix`` (e.g. ``"gmail."``).
#     Used to decide whether the orchestrator should skip its
#     blocking prefetch step: if the agent already exposes a tool to
#     fetch fresh data, prefetching duplicates work and adds 8s of
#     latency."""
#     if not agent_id or not tool_prefix:
#         return False
#     agent = get_agent(agent_id)
#     if not agent:
#         return False
#     return any(
#         isinstance(tool, str) and tool.startswith(tool_prefix)
#         for tool in agent.get("tools", [])
#     )


# def _iter_agents():
#     for agent in list_agents():
#         yield agent.get("id", ""), agent


# def _resolve_agent_for_intent(intent: Dict[str, Any], instance: str) -> Optional[str]:
#     """Resolve which agent should handle this intent (hardcoded + dynamic from Firestore).
#
#     Deprecated: prefer ``_resolve_agents_for_intents`` for multi-intent queries.
#     Kept for callers that expect a single agent id (e.g. group_consent flow).
#     """
#     if intent.get("is_runtime_status"):
#         return "runtime-status"
#     if intent["is_gross"] or intent["is_assault_related"]:
#         return "agent-morality"
#     if intent["is_correction"]:
#         return "agent-learning"
#     if intent["is_intimacy"]:
#         return "agent-intimacy"
#     if intent.get("is_rag"):
#         return "agent-knowledge-retriever"
#     if intent["is_drive"]:
#         return "manager-drive"
#     if intent["is_email"]:
#         return "manager-email"
#     if intent["is_calendar"]:
#         return "manager-calendar"
#     if intent["is_web_search"]:
#         return "manager-web"
#
#     rules = _get_routing_rules()
#     for rule in sorted(rules, key=lambda r: r.get("priority", 99)):
#         agent_id = rule.get("agent_id", "")
#         if intent.get(f"matched_{agent_id}"):
#             return agent_id
#
#     return None


# _AGENT_INTENT_FLAGS: List[Tuple[str, str]] = [
#     ("is_rag", "agent-knowledge-retriever"),
#     ("is_drive", "manager-drive"),
#     ("is_email", "manager-email"),
#     ("is_calendar", "manager-calendar"),
#     ("is_web_search", "manager-web"),
# ]


# _FILENAME_EXT = re.compile(
#     r"\.(pdf|docx|xlsx|txt|csv|md|rtf|odt)\b",
#     re.IGNORECASE,
# )


# def _import_rag_helpers():
#     """Import lazy do retriever para evitar import circular."""
#     from agent_orchestration.knowledge_retriever import _had_recent_indexing
#     return _had_recent_indexing


# def _has_filename_hint(text: str) -> bool:
#     """Detecta mencao explicita a um arquivo com extensao conhecida
#     na mensagem do usuario (ex: 'cdc-portugues-2013.pdf', 'relatorio.docx')."""
#     if not text:
#         return False
#     return bool(_FILENAME_EXT.search(text))


# def _resolve_agents_for_intents(
#     intent: Dict[str, Any],
#     instance: str,
#     masked_text: str = "",
#     scope_key: str = "",
# ) -> List[str]:
#     """Resolve ALL agents that should handle this intent in parallel.
#
#     Multi-intent: when a query triggers more than one intent flag
#     (e.g., 'quais meus ultimos 5 emails?' matches both is_email AND is_rag
#     because of the `?` marker), we want BOTH agents to execute in parallel.
#
#     Returns a deduped list of agent ids in stable order. Returns an empty
#     list when no specialist agent matches (caller decides whether to
#     fall back to the jennifier orchestrator).
#
#     Defense-in-depth (F4d.9): when a personal-intent agent is in the
#     list, the knowledge-retriever is excluded. Personal intents
#     (email/calendar/drive) already have tools that fetch fresh data;
#     running the retriever in parallel would inflate cost without
#     adding value (the personal agent can call knowledge.retrieve as
#     a tool when needed).
#
#     Filename override (30/07/2026): when a query cita um arquivo com
#     extensao conhecida E houve indexing recente para este scope
#     (phone ou group_jid), o retriever e preferido sobre Drive/RAG
#     heuristic. O user acabou de indexar esse arquivo, entao ele
#     DEVE estar no RAG, nao no Drive.
#     """
#     if intent.get("is_runtime_status"):
#         return ["runtime-status"]
#     if intent.get("is_gross") or intent.get("is_assault_related"):
#         return ["agent-morality"]
#     if intent.get("is_correction"):
#         return ["agent-learning"]
#     if intent.get("is_intimacy"):
#         return ["agent-intimacy"]
#
#     if (
#         masked_text
#         and scope_key
#         and _has_filename_hint(masked_text)
#         and _import_rag_helpers()(scope_key)
#     ):
#         return ["agent-knowledge-retriever"]
#
#
#
#     seen: set = set()
#     agents: List[str] = []
#     for flag, agent_id in _AGENT_INTENT_FLAGS:
#         if intent.get(flag) and agent_id not in seen:
#             agents.append(agent_id)
#             seen.add(agent_id)
#
#     personal_present = any(a.startswith("manager-") for a in agents)
#     if personal_present:
#         agents = [a for a in agents if a != "agent-knowledge-retriever"]
#
#     return agents


# PERSONAL_INTENTS = {"is_calendar", "is_drive", "is_email"}


# def _is_personal_intent(intent: Dict[str, Any]) -> bool:
#     """Check if intent involves personal data (calendar, email, drive)."""
#     return any(intent.get(k) for k in PERSONAL_INTENTS)


def _is_group_message(payload: Dict[str, Any]) -> bool:
    """Check if message is from a WhatsApp group."""
    extra = payload.get("extra", {})
    remote_jid = extra.get("remote_jid", payload.get("phone", ""))
    return "@g.us" in str(remote_jid)


def _extract_group_jid(payload: Dict[str, Any]) -> str:
    """Extract group JID from payload."""
    extra = payload.get("extra", {})
    remote_jid = extra.get("remote_jid", "")
    if "@g.us" in str(remote_jid):
        return remote_jid.split("@")[0] + "@g.us"
    return ""


async def _download_attachment_bytes(envelope: Dict[str, Any]) -> Optional[bytes]:
    """Baixa attachment do envelope: base64 inline primeiro, depois POST sob demanda."""
    extra = envelope.get("extra", {})
    doc_b64 = extra.get("doc_base64", "")
    if doc_b64:
        try:
            return base64.b64decode(doc_b64)
        except Exception as exc:
            logger.warning("attachment: base64 inline decode failed: %s", exc)
    message_id = envelope.get("message_id", "")
    instance = envelope.get("instance", "")
    remote_jid = extra.get("remote_jid", "")
    if message_id and instance:
        try:
            from core.evolution_client import get_base64_from_media_message
            result = await get_base64_from_media_message(
                instance=instance,
                message_id=message_id,
                remote_jid=remote_jid,
            )
            doc_b64 = result.get("base64", "")
            if doc_b64:
                return base64.b64decode(doc_b64)
        except Exception as exc:
            logger.warning("attachment: get_base64_from_media_message failed: %s", exc)
    return None


async def _extract_text_from_attachment(
    envelope: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    """Baixa attachment e extrai texto por mimetype.

    Returns:
        {"text": str, "source_name": str, "mimetype": str, "raw_size": int}
        ou None se falhar.
    """
    extra = envelope.get("extra", {})
    if not extra.get("has_document"):
        return None
    mimetype = (extra.get("doc_mimetype") or "").lower()
    source_name = extra.get("doc_file_name") or "document"
    raw = await _download_attachment_bytes(envelope)
    if raw is None:
        return None
    text = ""
    try:
        if "pdf" in mimetype:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(raw))
            text = "\n".join((page.extract_text() or "") for page in reader.pages)
        elif "wordprocessingml" in mimetype or "msword" in mimetype:
            from docx import Document
            doc = Document(io.BytesIO(raw))
            text = "\n".join(p.text for p in doc.paragraphs if p.text)
        elif "spreadsheetml" in mimetype or "excel" in mimetype:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            chunks = []
            for sheet in wb.worksheets:
                chunks.append(f"--- {sheet.title} ---")
                for row in sheet.iter_rows(values_only=True):
                    line = ",".join("" if v is None else str(v) for v in row)
                    chunks.append(line)
            text = "\n".join(chunks)
        else:
            try:
                text = raw.decode("utf-8", errors="replace")
            except Exception:
                text = ""
    except Exception as exc:
        logger.warning("attachment: text extract failed mimetype=%s err=%s", mimetype, exc)
        return None
    return {
        "text": text,
        "source_name": source_name,
        "mimetype": mimetype,
        "raw_size": len(raw),
    }


async def _persist_attachment(
    envelope: Dict[str, Any],
    extracted: Dict[str, Any],
    save_to_rag: bool,
) -> Dict[str, Any]:
    """Persiste attachment delegando ao knowledge router (Fase G).

    O router decide skill (MIME), escopo (individual/grupo) e destino
    (Firestore Vector por padrão; Google Drive apenas se o user pedir
    explicitamente). Para manter compatibilidade com a assinatura
    legada, ``save_to_rag=False`` força o caminho Drive.

    F4d.6+: Quando o destino é RAG (Firestore Vector), também roda o
    categorizer (LLM + heuristica) para preencher class/group/theme do
    documento. Sem isso os chunks sao indexados sem metadata de filtro,
    comprometendo a retrieval (Fase H).
    """
    from agent_orchestration.knowledge_router import route_attachment

    decision_input = envelope.copy()
    decision_input["_drive_extracted"] = extracted
    user_text = "memorizar" if save_to_rag else "drive"
    decision = await route_attachment(decision_input, user_text)
    skill = decision.get("skill")
    if skill is None:
        return decision.get("persist_result") or {
            "error": "no_skill",
            "detail": decision,
        }
    extracted_payload = decision.get("extracted") or extracted
    category_metadata = decision.get("category") or {}
    persist_scope = decision.get("scope", "private")

    if (
        skill is not None
        and save_to_rag
        and not category_metadata
        and persist_scope != "drive"
    ):
        try:
            from agent_orchestration.knowledge_router import (
                categorize_and_extract,
            )
            extraction = await categorize_and_extract(envelope, skill)
            if extraction.get("extracted"):
                extracted_payload = extraction["extracted"]
            if extraction.get("category"):
                category_metadata = extraction["category"]
                decision["extracted"] = extracted_payload
                decision["category"] = category_metadata
        except Exception as exc:
            logger.warning("categorize_and_extract failed in _persist_attachment: %s", exc)

    persist_result = await skill.persist(
        envelope,
        extracted_payload,
        persist_scope,
        metadata=category_metadata,
    )
    if persist_result.get("error"):
        return persist_result
    persist_result["scope"] = persist_scope
    persist_result["skill_name"] = decision.get("skill_name")
    if category_metadata:
        persist_result["category"] = category_metadata
    return persist_result


# def _prefetch_tone_guide(intent: Dict[str, Any]) -> str:
#     if intent.get("is_calendar"):
#         return (
#             "Formate estes dados como uma conversa com um amigo. "
#             "Exemplo: 'Sua agenda hoje está assim: de manhã você tem uma reunião às 10h, "
#             "depois está livre até as 15h. Quer que eu te lembre de algo? 📅'\n"
#         )
#     if intent.get("is_email"):
#         return (
#             "Formate estes dados como uma conversa com um amigo. "
#             "Para listas, use tabela ASCII em bloco ``` com colunas: "
#             "Remetente | Assunto | Data. "
#             "Exemplo: 'Achei 3 emails! 📧'\n"
#         )
#     return (
#         "Formate estes dados como uma conversa com um amigo. "
#         "Para listas de arquivos, use tabela ASCII em bloco ``` com colunas: "
#         "Nome | Tipo | Modificado. "
#         "Exemplo: 'Encontrei estes arquivos! 📁'\n"
#     )


def _prefetch_nickname(first_name: str) -> Optional[str]:
    """G7: Pre-resolve apelido do JSON estatico, sem LLM tool loop."""
    try:
        import json as _json
        data_file = os.path.join(
            os.path.dirname(__file__), "data", "nicknames.json"
        )
        with open(data_file, "r", encoding="utf-8") as f:
            data = _json.load(f)
        normalized = first_name.strip().title()
        nicknames = data.get(normalized, [])
        if nicknames:
            return nicknames[0]
        if not data.get("_comment"):
            # fallback: gera diminutivo
            if len(normalized) >= 8:
                return normalized[:4]
            elif len(normalized) >= 6:
                return normalized[:3]
            elif len(normalized) >= 3:
                return normalized[:2]
            else:
                return normalized
    except Exception:
        pass
    return None


def _generate_diminutive(name: str) -> str:
    """Fallback: gera diminutivo carinhoso do primeiro nome."""
    if len(name) <= 3:
        return name + name[-1]
    elif len(name) <= 6:
        return name[:2]
    else:
        return name[:4]


# def _is_read_query(text: str) -> bool:
#     """Fase B: detecta se query eh de leitura (pre-fetch) ou escrita (tool loop)."""
#     text_lower = text.lower()
#     write_kw = {"cria", "criar", "envia", "enviar", "manda", "mandar", "agenda", "agendar",
#                 "marca", "marcar", "deleta", "deletar", "apaga", "apagar", "atualiza", "atualizar",
#                 "remove", "remover", "sobe", "upload"}
#     if any(kw in text_lower for kw in write_kw):
#         return False
#     return True


def _extract_search_terms(text: str) -> str:
    """Extrai termos relevantes da query do usuario para busca no Drive."""
    stopwords = {"jen", "jennifer", "voce", "pode", "me", "meu", "minha", "meus",
                 "minhas", "um", "uma", "para", "com", "que", "nao", "sim", "por",
                 "favor", "obrigado", "acessar", "consegue", "veja", "traga", "acesse",
                 "busca", "buscar", "encontra", "encontrar", "ache", "achar",
                 "listar", "lista", "mostre", "mostrar", "quero", "preciso",
                 "de", "da", "do", "no", "na", "mais", "a", "o", "e", "se", "ja"}
    words = text.lower().split()
    terms = [w.strip(",.!?;:") for w in words if w.strip(",.!?;:") not in stopwords and len(w.strip(",.!?;:")) > 2]
    return " ".join(terms[:10])


async def _prefetch_calendar(phone: str, instance: str = "") -> Optional[str]:
    """Fase B: pre-busca eventos do dia sem LLM."""
    try:
        from tools.google_calendar import list_events
        from datetime import datetime, timezone, timedelta
        brt = timezone(timedelta(hours=-3))
        hoje = datetime.now(brt)
        result = await list_events(
            phone,
            time_min=hoje.strftime("%Y-%m-%dT00:00:00-03:00"),
            time_max=hoje.strftime("%Y-%m-%dT23:59:59-03:00"),
            max_results=50,
            instance=instance,
        )
        if isinstance(result, str):
            try:
                result = json.loads(result)
            except Exception:
                result = {}
        if not isinstance(result, dict):
            return None
        events = result.get("events", [])
        if not events:
            return None
        return json.dumps(events, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Prefetch calendar failed: {e}")
        return None


async def _prefetch_email(phone: str, instance: str = "") -> Optional[str]:
    """Fase B: pre-busca ultimos emails sem LLM."""
    try:
        from tools.google_gmail import search_messages
        result = await search_messages(
            phone,
            "in:inbox newer_than:30d",
            max_results=10,
            instance=instance,
        )
        if isinstance(result, str):
            try:
                result = json.loads(result)
            except Exception:
                result = {}
        if not isinstance(result, dict):
            return None
        messages = result.get("messages", [])
        if not messages:
            return None
        return json.dumps(messages, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Prefetch email failed: {e}")
        return None


async def _prefetch_drive(phone: str, query_text: str = "", instance: str = "") -> Optional[str]:
    """Fase B: pre-busca arquivos no Drive sem LLM."""
    try:
        from tools.google_drive import search_files
        result = await search_files(phone, query_text or "", max_results=20, instance=instance)
        files = result.get("files", [])
        if not files:
            return None
        # FIX (15/08/2026): aplicar o filtro de curriculo_padrao tambem no
        # prefetch para manter consistencia com a tool search_drive_files
        # (do DeepAgent). Sem isso, o prefetch retorna 3 arquivos enquanto a
        # tool prioriza 1, gerando inconsistencias no prompt do LLM.
        try:
            from tools.memory import get_fact_by_key
            default_filename = await get_fact_by_key("curriculo_padrao", phone)
            if default_filename:
                query_norm = (query_text or "").strip().lower()
                if any(kw in query_norm for kw in ("curriculo", "currículo", "resum", "cv ")):
                    prioritized, others = [], []
                    default_norm = default_filename.strip().lower()
                    for f in files:
                        fname = str(f.get("name") or "").strip().lower()
                        if fname == default_norm:
                            prioritized.append(f)
                        else:
                            others.append(f)
                    if prioritized:
                        files = prioritized + others
        except Exception as exc:
            logger.debug("prefetch_curriculo_filter_skipped: %s", exc)
        return json.dumps(files, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Prefetch drive failed: {e}")
        return None


async def _prefetch_drive_multi(phone: str, text: str, instance: str = "") -> Optional[str]:
    """D1: 3 queries paralelas no Drive, usa a com mais resultados."""
    query1 = _extract_search_terms(text)
    query2 = " ".join(w for w in text.lower().split()
                      if len(w.strip(",.!?;:")) > 3)[:5]
    results: List[Any] = []
    try:
        results = await asyncio.wait_for(
            asyncio.gather(
                _prefetch_drive(phone, query1, instance),
                _prefetch_drive(phone, query2, instance),
                _prefetch_drive_docs(phone, query1, instance),
                _prefetch_drive_docs(phone, query2, instance),
                return_exceptions=True,
            ),
            timeout=PREFETCH_DRIVE_MULTI_TIMEOUT_SEC,
        )
    except asyncio.TimeoutError:
        logger.warning(
            "prefetch_drive_multi_timeout timeout_sec=%s",
            PREFETCH_DRIVE_MULTI_TIMEOUT_SEC,
        )
    best, best_count = None, 0
    for r in results:
        if isinstance(r, str):
            try:
                data = json.loads(r)
                if isinstance(data, list) and len(data) > best_count:
                    best, best_count = r, len(data)
            except Exception:
                pass
    return best


async def _prefetch_drive_docs(phone: str, query_text: str = "", instance: str = "") -> Optional[str]:
    """Prefetch so documentos e apresentacoes do Drive (atas, slides)."""
    try:
        from tools.google_drive import search_files
        query_parts = []
        if query_text:
            query_parts.append(f"name contains '{query_text}'")
        query_parts.append("mimeType contains 'document' or mimeType contains 'presentation'")
        result = await search_files(
            phone,
            query_text or "",
            mime_type="application/vnd.google-apps.document",
            max_results=20,
            instance=instance,
        )
        alt = await search_files(
            phone,
            query_text or "",
            mime_type="application/vnd.google-apps.presentation",
            max_results=20,
            instance=instance,
        )
        all_files = result.get("files", []) + alt.get("files", [])
        if not all_files:
            return None
        seen = set()
        unique = []
        for f in all_files:
            if f["id"] not in seen:
                seen.add(f["id"])
                unique.append(f)
        return json.dumps(unique[:20], ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Prefetch drive docs failed: {e}")
        return None


def _has_real_data(prefetch_text: str) -> bool:
    """A1: Retorna True so se prefetch trouxe dados reais (nao mensagens de vazio)."""
    empty_signals = {"nenhum", "não encontrou", "nao encontrou", "sem resultados",
                     "0 resultados", "error", "failed"}
    text_lower = prefetch_text.lower()
    return not any(m in text_lower for m in empty_signals)

ACCEPTANCE_KEYWORDS = {"sim", "pode", "ok", "claro", "aceito", "perfeito", "otimo", "pode sim"}
REJECTION_KEYWORDS = {"nao", "não", "prefiro nao", "prefiro não", "recuso", "deixa"}


def _short_confirmation(text: str) -> Optional[bool]:
    normalized = _normalize_text(text).strip(" .,!?:;")
    if normalized in {_normalize_text(value) for value in ACCEPTANCE_KEYWORDS}:
        return True
    if normalized in {_normalize_text(value) for value in REJECTION_KEYWORDS}:
        return False
    return None


def _get_db():
    """Get Firestore client (fallback robusto)."""
    try:
        from google.cloud import firestore
        import os
        project = os.getenv("GCP_PROJECT") or os.getenv("GCLOUD_PROJECT")
        if not project:
            return None
        return firestore.Client(project=project)
    except Exception:
        return None


def _get_conversation_history(phone: str, limit: int = 10) -> str:
    """A2: Busca historico de conversas do Firestore para injetar no prompt."""
    db = _get_db()
    if not db:
        return ""
    try:
        import hashlib
        ph = hashlib.sha256(phone.encode()).hexdigest()[:16]
        docs = (
            db.collection("contatos").document(ph)
            .collection("historico")
            .order_by("ts", direction="DESCENDING")
            .limit(limit)
            .stream()
        )
        msgs = []
        for d in docs:
            data = d.to_dict()
            direction = data.get("direction", "in")
            text = data.get("text", "")[:80]
            prefix = "Usuario" if direction == "in" else "Jennifer"
            msgs.append(f"{prefix}: {text}")
        return "\n".join(reversed(msgs)) if msgs else ""
    except Exception as e:
        logger.warning(f"Conversation history failed: {e}")
        return ""


async def _get_context_for_prompt(phone: str, limit: int = 10) -> str:
    try:
        from core.rag import get_conversation_history

        result = await get_conversation_history(phone, limit)
        if result:
            return result
    except Exception:
        pass
    try:
        return _get_conversation_history(phone, limit)
    except Exception:
        return ""


async def _search_memory(phone: str, query: str, limit: int = 5) -> str:
    try:
        from core.rag import search_conversation_memory

        results = await search_conversation_memory(phone, query, limit)
        memories = []
        for result in results:
            text = result.get("text", "")[:300]
            direction = result.get("direction", "in")
            identity = "Usuario" if direction == "in" else result.get("response_identity", "Jennifer")
            memories.append(f"- {identity}: {text}")
        return "\n".join(memories)
    except Exception as exc:
        logger.warning("RAG search failed: %s", exc)
        return ""


async def _index_message(phone: str, text: str, direction: str, **metadata: Any) -> Dict[str, Any]:
    try:
        from core.rag import index_conversation_message

        result = await index_conversation_message(
            phone=phone,
            text=text,
            direction=direction,
            message_id=metadata.get("message_id"),
            conversation_id=metadata.get("conversation_id"),
            turn_id=metadata.get("turn_id"),
            agent_id=metadata.get("agent_id"),
            response_identity=metadata.get("response_identity", "Jennifer"),
        )
        if result.get("status") == "indexed":
            logger.info("RAG message indexed: direction=%s doc_id=%s", direction, result.get("doc_id"))
        else:
            logger.warning("RAG message not indexed: direction=%s reason=%s", direction, result.get("reason"))
        return result
    except Exception as exc:
        logger.error("RAG message indexing failed: %s", exc)
        return {"status": "error", "reason": str(exc)}


async def index_audio_failure_for_audit(body: Dict[str, Any], error_code: str) -> Dict[str, Any]:
    try:
        phone = body.get("phone", "")
        message_id = body.get("message_id") or (body.get("extra") or {}).get("message_id") or ""
        if not phone:
            return {"status": "skipped", "reason": "missing_phone"}
        timestamp_brt = now_brt().isoformat(timespec="minutes")
        marker_text = mask_pii(
            f"[audio transcription failed at {timestamp_brt} reason={error_code}]"
        )
        return await _index_message(
            phone,
            marker_text,
            "in",
            message_id=f"audio-fail:{message_id}" if message_id else None,
            conversation_id=(body.get("extra") or {}).get("remote_jid") or phone,
            turn_id=message_id,
            agent_id="audio-transcriber",
            response_identity="AudioAudit",
        )
    except Exception as exc:
        logger.warning("audio failure audit indexing failed: %s", type(exc).__name__)
        return {"status": "error", "reason": type(exc).__name__}


# async def _execute_single_specialist(
#     specialist_id: str,
#     intent: Dict[str, Any],
#     payload: Dict[str, Any],
#     masked_text: str,
#     extra: Dict[str, Any],
#     instance: str,
#     phone: str,
#     sender_name: str,
#     path: List[Dict[str, Any]],
# ) -> Dict[str, Any]:
#     """Run one specialist agent end-to-end (legacy single-agent path).
#
#     Wraps the access_guardian check, prefetch logic and tool execution
#     that used to live inline in `orchestrate`. Returns the agent's
#     result dict (ready for ``_finalize_orchestration``).
#     """
#     agent = get_agent(specialist_id)
#     if not agent or not agent.get("enabled", True):
#         if not agent:
#             try:
#                 from agent_loader import _load_all
#                 _load_all()
#                 agent = get_agent(specialist_id)
#             except Exception as exc:
#                 logger.warning("emergency_reload_failed err=%s", exc)
#     if not agent or not agent.get("enabled", True):
#         path.append({"step": 2, "phase": "fallback_to_orchestrator", "reason": "specialist_disabled"})
#         logger.warning(
#             "specialist_agent_missing specialist_id=%s falling_back_to_orchestrator",
#             specialist_id,
#         )
#         orchestrator_id = await _get_orchestrator(instance)
#         if not orchestrator_id:
#             return _error_response(503, "no_orchestrator", "Nenhum orchestrator disponivel")
#         orchestrator = get_agent(orchestrator_id)
#         if not orchestrator:
#             return _error_response(503, "agent_not_found", f"Orchestrator {orchestrator_id} nao encontrado")
#         return await _execute_agent(orchestrator, masked_text, payload, extra)
#
#     agent_copy = dict(agent)
#     prefetch_data = None
#
#     skip_guard = not _is_personal_intent(intent) and (
#         specialist_id.startswith("manager-")
#         and _agent_has_tool(specialist_id, "calendar.")
#         or specialist_id.startswith("manager-")
#         and _agent_has_tool(specialist_id, "gmail.")
#         or specialist_id.startswith("manager-")
#         and _agent_has_tool(specialist_id, "drive.")
#     )
#
#     if skip_guard:
#         guard_result = {"verdict": "noop", "decision": {}, "prefetch": None, "trace": ["guard_skipped:manager_has_tools"]}
#     else:
#         guard_result = await _run_guard_graph(payload, masked_text, intent)
#     guard_verdict = (guard_result or {}).get("verdict", "noop")
#     if guard_verdict in {"deny", "request_oauth"}:
#         decision = (guard_result or {}).get("decision") or {}
#         link = decision.get("oauth_link", "")
#         if guard_verdict == "request_oauth" and link:
#             reply = (
#                 "Oi! Para acessar " + decision.get("capability", "essa ferramenta") +
#                 ", preciso que voce autorize sua conta Google. "
#                 f"Acesse este link e faca o login: {link}"
#             )
#         else:
#             reply = (
#                 "Oi! Essa acao so pode ser executada pelo proprietario "
#                 "da conta WhatsApp."
#             )
#         return {
#             "reply": reply,
#             "delay_ms": 0,
#             "presence": "composing",
#             "metadata": {
#                 "agent_id": "access_guardian",
#                 "guardian_verdict": guard_verdict,
#                 "guardian_reason": decision.get("reason", ""),
#                 "guardian_capability": decision.get("capability", ""),
#                 "response_identity": "Jennifer",
#                 "blocked": True,
#             },
#         }
#
#     if guard_verdict == "allow" and _is_personal_intent(intent):
#         try:
#             ack_map = {
#                 "calendar": "Só um instante. Vou ver sua agenda... 📅",
#                 "drive": "Só um instante. Vou procurar aqui... 📁",
#                 "email": "Só um instante. Vou buscar seus emails... 📧",
#             }
#             ack_intent = "calendar" if intent.get("is_calendar") else (
#                 "drive" if intent.get("is_drive") else "email"
#             )
#             ack_text = ack_map.get(ack_intent, "Só um instante... ⏳")
#             from core.evolution_client import send_presence, send_text
#             ack_delay_ms = max(1500, calculate_delay_ms(ack_text))
#             asyncio.create_task(send_presence(instance, phone, "composing", remote_jid=extra.get("remote_jid", "")))
#             asyncio.create_task(send_text(
#                 instance=instance, phone=phone, text=ack_text,
#                 delay_ms=ack_delay_ms, presence="composing",
#                 remote_jid=extra.get("remote_jid", ""),
#             ))
#         except Exception:
#             pass
#
#     if guard_result.get("prefetch"):
#         prefetch_data = guard_result["prefetch"]
#
#     if prefetch_data is None and _is_read_query(masked_text):
#         if _agent_has_tool(specialist_id, "calendar."):
#             pass
#         elif _agent_has_tool(specialist_id, "gmail."):
#             pass
#         elif _agent_has_tool(specialist_id, "drive."):
#             pass
#         else:
#             try:
#                 if intent.get("is_calendar"):
#                     prefetch_data = await asyncio.wait_for(
#                         _prefetch_calendar(phone, instance), timeout=4)
#                 elif intent.get("is_email"):
#                     prefetch_data = await asyncio.wait_for(
#                         _prefetch_email(phone, instance), timeout=4)
#                 elif intent.get("is_drive"):
#                     prefetch_data = await asyncio.wait_for(
#                         _prefetch_drive_multi(phone, masked_text, instance), timeout=4)
#             except asyncio.TimeoutError:
#                 logger.warning(f"Prefetch timeout for {specialist_id}")
#                 prefetch_data = None
#             except Exception as e:
#                 logger.warning(f"Prefetch failed for {specialist_id}: {e}")
#                 prefetch_data = None
#
#     if prefetch_data and _has_real_data(prefetch_data):
#         prefetch_data = mask_pii(prefetch_data)
#         data_label = "CALENDARIO" if intent.get("is_calendar") else \
#                      "EMAILS" if intent.get("is_email") else "DRIVE"
#         tone_guide = _prefetch_tone_guide(intent)
#         agent_copy["system_prompt"] += (
#             f"\n\n[DADOS PRE-CARREGADOS DO {data_label}]\n{prefetch_data}\n\n"
#             f"{tone_guide}"
#             "NAO chame ferramentas — os dados ja estao prontos."
#         )
#         agent_copy["tools"] = []
#
#     if _is_group_message(payload) and intent.get("is_drive"):
#         group_jid = _extract_group_jid(payload)
#         if group_jid:
#             try:
#                 from tools.group import get_group_drive_folder, get_group_info
#                 drive_folder = await get_group_drive_folder(group_jid)
#                 group_info = await get_group_info(group_jid)
#                 group_name = group_info.get("name", "grupo")
#                 if drive_folder:
#                     agent_copy["system_prompt"] += (
#                         f"\n\n[CONTEXTO DE GRUPO]\n"
#                         f"Voce esta respondendo no grupo '{group_name}'. "
#                         f"Use APENAS a pasta do grupo (ID: {drive_folder}) para buscas. "
#                         f"Nao acesse outras pastas do Drive do owner."
#                     )
#                 else:
#                     agent_copy["system_prompt"] += (
#                         f"\n\n[CONTEXTO DE GRUPO]\n"
#                         f"Voce esta respondendo no grupo '{group_name}'. "
#                         f"Este grupo ainda nao tem uma pasta no Drive associada. "
#                         f"Informe o usuario e pergunte se quer criar uma."
#                     )
#             except Exception:
#                 pass
#
#     path.append({"step": 2, "phase": "specialist", "agent": specialist_id,
#                  "prefetch": bool(prefetch_data),
#                  "reason": {k: v for k, v in intent.items() if v}})
#     with current_tracker().stage("execute_agent", agent_id=specialist_id):
#         return await _execute_agent(agent_copy, masked_text, payload, extra)


# async def _execute_multi_specialists_parallel(
#     specialist_ids: List[str],
#     intent: Dict[str, Any],
#     payload: Dict[str, Any],
#     masked_text: str,
#     extra: Dict[str, Any],
#     instance: str,
#     phone: str,
#     sender_name: str,
#     path: List[Dict[str, Any]],
# ) -> Dict[str, Any]:
#     """Run multiple specialist agents in parallel and merge results.
#
#     Multi-intent example: 'quais meus ultimos 5 emails?' matches both
#     ``is_email`` and ``is_rag`` (because of the `?` marker). The user
#     expects BOTH agents to execute and return combined output.
#
#     Behavior:
#     - Each agent is called via ``_execute_agent`` concurrently via
#       ``asyncio.gather``.
#     - Per-agent errors are isolated (return_exceptions=True); one
#       failing agent doesn't break the others.
#     - The final reply is the concatenation of all non-empty agent
#       outputs, each prefixed with the agent id for traceability.
#     """
#     path.append({
#         "step": 2,
#         "phase": "multi_specialist_parallel",
#         "agents": list(specialist_ids),
#         "reason": {k: v for k, v in intent.items() if v},
#     })
#
#     async def _run_one(specialist_id: str) -> Tuple[str, Optional[Dict[str, Any]]]:
#         try:
#             agent = get_agent(specialist_id)
#             if not agent or not agent.get("enabled", True):
#                 return specialist_id, None
#             agent_copy = dict(agent)
#             res = await _execute_agent(agent_copy, masked_text, payload, extra)
#             return specialist_id, res
#         except Exception as exc:
#             logger.exception("multi_specialist_failed agent=%s", specialist_id)
#             return specialist_id, {"reply": "", "metadata": {"error": type(exc).__name__}}
#
#     async def _run_one_safe(specialist_id: str) -> Tuple[str, Optional[Dict[str, Any]]]:
#         try:
#             return await asyncio.wait_for(
#                 _run_one(specialist_id),
#                 timeout=MULTI_SPECIALIST_TIMEOUT_SEC,
#             )
#         except asyncio.TimeoutError:
#             logger.warning(
#                 "multi_specialist_task_timeout agent=%s timeout_sec=%s",
#                 specialist_id, MULTI_SPECIALIST_TIMEOUT_SEC,
#             )
#             path.append({
#                 "step": 2,
#                 "phase": "multi_specialist_task_timeout",
#                 "agent_id": specialist_id,
#                 "timeout_sec": MULTI_SPECIALIST_TIMEOUT_SEC,
#             })
#             return specialist_id, {
#                 "reply": "",
#                 "metadata": {"error": "agent_timeout", "agent_id": specialist_id},
#             }
#
#     pairs = await asyncio.gather(*[_run_one_safe(s) for s in specialist_ids])
#     successful = [(sid, r) for sid, r in pairs if r and r.get("reply")]
#
#     if not successful:
#         return _error_response(500, "multi_agent_empty",
#                                "Os agentes nao conseguiram gerar uma resposta.")
#
#     if len(successful) == 1:
#         sid, r = successful[0]
#         return r
#
#     sections: List[str] = []
#     for sid, r in successful:
#         reply = (r.get("reply") or "").strip()
#         if not reply:
#             continue
#         sections.append(reply)
#     if not sections:
#         return _error_response(500, "multi_agent_empty",
#                                "Os agentes nao conseguiram gerar uma resposta.")
#
#     merged_reply = "\n\n---\n\n".join(sections)
#     merged_metadata = {
#         "agent_id": "+".join(sid for sid, _ in successful),
#         "response_identity": "Jennifer",
#         "multi_agent": True,
#         "agents_executed": [sid for sid, _ in successful],
#         "agent_errors": [
#             {"agent_id": sid, "error": (r or {}).get("metadata", {}).get("error")}
#             for sid, r in pairs if r and (r.get("metadata") or {}).get("error")
#         ],
#     }
#     primary_delay = max(
#         (r.get("delay_ms") or 0 for _, r in successful),
#         default=0,
#     )
#     return {
#         "reply": merged_reply,
#         "delay_ms": primary_delay,
#         "presence": "composing",
#         "metadata": merged_metadata,
#     }


def _detect_runtime_status(text: str) -> bool:
    keywords = (
        "quantos agentes", "quais agentes", "agentes funcionando", "agentes ativos",
        "agentes rodando", "status dos agentes", "o que esta rodando", "o que está rodando",
        "listar agentes", "liste os agentes",
    )
    return any(kw in text.lower() for kw in keywords)


def _detect_intimacy(text: str) -> bool:
    keywords = (
        "me chame de", "pode me chamar de", "meu apelido",
        "meu nome é", "meu nome e", "como devo te chamar",
    )
    return any(kw in text.lower() for kw in keywords)


def _detect_correction(text: str) -> bool:
    keywords = ("na verdade", "não é assim", "nao e assim", "errado", "errada")
    return any(kw in text.lower() for kw in keywords)


def _detect_morality(text: str) -> bool:
    gross = (
        "puta", "merda", "caralho", "fdp", "porra",
        "buceta", "viado", "bicha", "desgraça",
        "foder", "fode", "piranha", "vagabunda", "puto",
        "bosta", "porcaria", "desgraçado",
    )
    assault = ("assedio", "abuso", "estupro", "violencia", "agressao",
               "ameaça", "ameaca", "chantagem")
    t = text.lower()
    return any(kw in t for kw in gross) or any(kw in t for kw in assault)


def _detect_web(text: str) -> bool:
    keywords = (
        "pesquisar", "buscar na internet", "busque na internet", "procure na web",
        "pesquise na web", "noticia atual", "noticias atuais", "pesquisa sobre",
    )
    t = text.lower()
    if any(kw in t for kw in keywords):
        return True
    import re
    if re.search(r"https?://\S+", text, flags=re.IGNORECASE):
        return True
    return False


async def _handle_runtime_status(payload: dict, instance: str, phone: str) -> dict:
    from core.agent_status import build_agent_inventory, format_inventory_reply
    inventory = build_agent_inventory(instance=instance, phone=phone)
    reply = format_inventory_reply(inventory)
    return {
        "reply": reply,
        "delay_ms": calculate_delay_ms(reply),
        "presence": calculate_presence(),
        "metadata": {
            "agent_id": "runtime-status",
            "route": "deterministic",
            "response_identity": "Jennifer",
            "counts": inventory["counts"],
            "generated_at": inventory["generated_at"],
        },
    }


async def _handle_morality(payload: dict, masked_text: str, sender_name: str,
                           cache_key, instance: str, phone: str) -> dict:
    path = [{"step": 1, "phase": "morality_handler"}]
    agent = get_agent("agent-morality")
    if agent:
        result = await _execute_agent(dict(agent), masked_text, payload, payload.get("extra", {}))
        return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)
    result = {
        "reply": "Mensagem bloqueada por violar as politicas de respeito.",
        "delay_ms": 0,
        "presence": "composing",
        "metadata": {"agent_id": "morality-guard", "blocked": True},
    }
    return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)


async def _handle_correction(payload: dict, masked_text: str, sender_name: str,
                              cache_key) -> dict:
    path = [{"step": 1, "phase": "correction_handler"}]
    agent = get_agent("agent-learning")
    if agent:
        result = await _execute_agent(dict(agent), masked_text, payload, payload.get("extra", {}))
        return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)
    result = {
        "reply": "Obrigado pela correcao! Vou aprender com isso.",
        "delay_ms": 0,
        "presence": "composing",
        "metadata": {"agent_id": "correction-handler"},
    }
    return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)


async def _handle_intimacy(payload: dict, masked_text: str, sender_name: str,
                            cache_key, first_name: str, phone: str) -> dict:
    path = [{"step": 1, "phase": "intimacy_handler"}]
    agent = get_agent("agent-intimacy")
    if agent:
        agent_copy = dict(agent)
        if first_name and not has_nickname(phone):
            suggested = _prefetch_nickname(first_name)
            if not suggested:
                suggested = _generate_diminutive(first_name)
            agent_copy["system_prompt"] = agent_copy.get("system_prompt", "") + (
                f"\n\n[CONTEXTO DE INTIMIDADE]\n"
                f"Primeiro nome do usuario: {first_name}. Apelido sugerido: {suggested}\n"
            )
        result = await _execute_agent(agent_copy, masked_text, payload, payload.get("extra", {}))
        return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)
    result = {
        "reply": f"Oi {first_name}! Como posso te chamar?",
        "delay_ms": 0,
        "presence": "composing",
        "metadata": {"agent_id": "intimacy-handler"},
    }
    return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)


async def _handle_web(payload: dict, masked_text: str, sender_name: str,
                       cache_key) -> dict:
    path = [{"step": 1, "phase": "web_handler"}]
    agent = get_agent("manager-web")
    if agent:
        result = await _execute_agent(dict(agent), masked_text, payload, payload.get("extra", {}))
        return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)
    return {
        "reply": "Funcionalidade de busca web indisponivel no momento.",
        "delay_ms": 0,
        "presence": "composing",
        "metadata": {"agent_id": "web-handler", "error": "agent_not_found"},
    }


def _merge_pipeline_results(results: list) -> dict:
    replies = [r["reply"] for r in results if r and r.get("reply")]
    delays = [r.get("delay_ms", 0) for r in results if r]
    return {
        "reply": "\n\n---\n\n".join(replies),
        "delay_ms": max(delays) if delays else 0,
        "presence": "composing",
        "metadata": {"multi_intent": True, "pipelines": len(replies)},
    }


_KNOWLEDGE_INTENTS = frozenset({"juridicas", "editais", "academica", "anotacoes"})

_CLASSIFIER_PROMPT = (
    "Classifique em UMA palavra:\n"
    "juridicas    - leis, codigos, artigos, decretos, normas, jurisprudencia\n"
    "editais      - licitacoes, concursos, pregoes, editais publicos\n"
    "academica    - teses, dissertacoes, artigos cientificos, papers\n"
    "anotacoes    - lembretes e notas pessoais SEM horario nem agenda (memoria)\n"
    "ferramentas  - agenda/calendario (criar, listar, marcar evento, compromisso, reuniao com horario/data), email, drive, pesquisa web\n"
    "conversa     - saudacoes, ajuda, perguntas genericas, buscas na base de conhecimento\n\n"
    "Pergunta: {text}\n\n"
    "Categoria:"
)


async def _classify_intent_llm(text: str) -> str:
    from langchain_openai import ChatOpenAI
    from core.secrets import get_secret

    valid = {"juridicas", "editais", "academica", "anotacoes", "ferramentas", "conversa"}
    # FIX (15/08/2026): garantir que o texto com acentuacao PT-BR nao quebre
    # o encode ASCII no cliente httpx/OpenAI para Groq. Forcar ASCII-safe
    # via transliteracao (unidecode fallback para caracteres que httpx
    # nao consegue serializar em UTF-8).
    import unicodedata
    text_nfkd = unicodedata.normalize("NFKD", text[:500])
    text_ascii_safe = text_nfkd.encode("ascii", errors="ignore").decode("ascii")
    prompt = _CLASSIFIER_PROMPT.format(text=text_ascii_safe)

    # Classificador unico: DeepSeek V4 Flash.
    # FIX (17/08/2026): bloco Groq removido — falhava sistematicamente com
    # "'ascii' codec can't encode characters" (erro do runtime Groq, nao do
    # texto, ja sanitizado). Cascade Groq->NVIDIA->DeepSeek removida em favor
    # de 1 unica chamada deterministica (menos latencia, menos custo).
    try:
        base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com/v1")
        deepseek_key = get_secret("DEEPSEEK_API_KEY") or os.getenv("DEEPSEEK_API_KEY", "")
        llm = ChatOpenAI(
            model="deepseek-v4-flash",
            api_key=deepseek_key,
            base_url=base_url,
            temperature=0,
            max_tokens=5,
            timeout=5,
            extra_body={"cache_mode": "default"},
        )
        result = await asyncio.to_thread(llm.invoke, prompt)
        raw = getattr(result, "content", str(result)).strip().lower()
        if raw in valid:
            return raw
    except Exception as exc:
        logger.warning("deepseek_intent_classifier_failed: %s", exc)

    return "conversa"


async def _setup_nickname_consent(payload: dict, first_name: str, phone: str) -> None:
    if not first_name or has_nickname(phone):
        return
    suggested = _prefetch_nickname(first_name)
    if not suggested:
        suggested = _generate_diminutive(first_name)
    from core.pending_actions import set_pending_action
    await set_pending_action(
        phone, "nickname_consent",
        {"first_name": first_name, "nickname": suggested},
    )


async def orchestrate(payload: Dict[str, Any]) -> Dict[str, Any]:
    _orchestrate_started = time.monotonic()
    instance = payload.get("instance", "jennifer")
    phone = payload.get("phone", "")
    text = payload.get("text", "")
    sender_name = payload.get("sender_name", "user")
    extra = payload.get("extra", {}) or {}

    try:
        result = await _orchestrate_inner(payload, instance, phone, text, sender_name, extra)
    finally:
        _elapsed = round((time.monotonic() - _orchestrate_started) * 1000, 2)
        logger.info(
            "orchestration_timing instance=%s phone=%s total_ms=%d",
            instance, phone, _elapsed,
        )
        try:
            from core.flood_protection import record_usage_metrics
            tracker = current_tracker()
            costs = tracker.costs() if tracker else {}
            remote_jid = payload.get("remote_jid") or extra.get("remote_jid", "")
            record_usage_metrics(phone=phone, group_id=remote_jid, instance=instance, costs=costs)
        except Exception as exc:
            logger.debug("record_usage_metrics in orchestrate finally failed: %s", exc)
    return result


# GUARDRAIL (17/08/2026): mapeamento keyword -> toolkit slug para auto-discovery.
# Toolkits com manager dedicado (1 API = 1 manager, §0.8). Calendar/gmail/
# drive sao roteados pelo TIER 1.7 deterministico (pipelines com guard).
# people/tasks/maps tem manager dedicado -> passam pelo DynamicManagerFactory.
_KEYWORD_TO_TOOLKIT: Dict[str, str] = {
    "linkedin": "linkedin",
    "youtube": "youtube",
    "github": "github",
    "notion": "notion",
    "onedrive": "onedrive",
    "google docs": "googledocs",
    "googledocs": "googledocs",
    "google sheets": "googlesheets",
    "googlesheets": "googlesheets",
    "sheets": "googlesheets",
    "google meet": "googlemeet",
    "googlemeet": "googlemeet",
    "meet": "googlemeet",
    "microsoft teams": "microsoft_teams",
    "teams": "microsoft_teams",
    "contatos": "people",
    "people": "people",
    "tarefas": "tasks",
    "tasks": "tasks",
    "maps": "maps",
    "rota": "maps",
    "mapa": "maps",
}


def _detect_dynamic_toolkit(text: str) -> Optional[str]:
    """Detecta toolkit slug via keywords. Retorna None se nenhuma match."""
    text_lower = text.lower()
    for keyword, slug in _KEYWORD_TO_TOOLKIT.items():
        if keyword in text_lower and api_registry.is_allowed(slug):
            return slug
    return None


async def _orchestrate_inner(payload: Dict[str, Any], instance: str, phone: str,
                             text: str, sender_name: str, extra: Dict[str, Any]) -> Dict[str, Any]:
    _tracker = new_tracker()
    set_current_tracker(_tracker)
    _tracker.add_costs(deepseek_input_tokens=0, deepseek_output_tokens=0)

    # ========================
    # ANTI-FLOOD / ANTI-DDOS / FINOPS SHIELD
    # ========================
    from core.flood_protection import check_and_record_message
    remote_jid = payload.get("remote_jid") or extra.get("remote_jid", "")
    group_name = extra.get("group_name", "")
    is_blocked, flood_details = check_and_record_message(
        phone=phone,
        group_id=remote_jid,
        instance=instance,
        text=text,
    )
    if is_blocked:
        if flood_details.get("quarantined"):
            from core.admin_notify import notify_admin_flood_alert
            asyncio.create_task(
                notify_admin_flood_alert(
                    phone=phone,
                    sender_name=sender_name or phone,
                    group_name=group_name or remote_jid,
                    burst_count=flood_details.get("burst_count", 0),
                    cost_usd=flood_details.get("estimated_cost_usd", 0.0),
                    cost_brl=flood_details.get("estimated_cost_brl", 0.0),
                    instance=instance,
                    text_preview=text,
                )
            )
        logger.warning("Message suppressed by flood protection phone=%s group=%s", phone, remote_jid)
        return {
            "reply": "",
            "presence": "paused",
            "metadata": {
                "quarantined": True,
                "error": "flood_protection_triggered",
            },
        }

    cache_key = _idempotency_key(payload)
    if cache_key and cache_key in _response_cache:
        cached = _response_cache[cache_key]
        if int(time.time()) - cached.get("ts", 0) < CACHE_TTL_SEC:
            response = copy.deepcopy(cached)
            response.pop("ts", None)
            response.setdefault("metadata", {})["cached"] = True
            return response

    first_name = _extract_first_name(sender_name)
    payload["first_name"] = first_name
    from agent_loader import ensure_user_registered
    ensure_user_registered(phone, sender_name=sender_name, instance=instance)
    masked_text = mask_pii(text)
    confirmation = _short_confirmation(masked_text)

    chat_context = await _get_context_for_prompt(phone, limit=10)
    payload["chat_context"] = chat_context


    # ========================
    # PENDING ACTIONS (pre-routing)
    # ========================
    from core.pending_actions import consume_pending_action, get_pending_action
    pending_action = await get_pending_action(phone)

    if pending_action and pending_action.get("action_type") == "attachment_mode":
        is_save = any(
            kw in masked_text.lower()
            for kw in ("memorize", "memorizar", "guarde", "guardar",
                       "indexe", "indexar", "armazene", "armazenar")
        )
        is_file = any(kw in masked_text.lower() for kw in ("salve", "salvar", "salva", "guarda"))
        if not (is_save or is_file):
            reply = "Responda apenas 'memorizar' para indexar no conhecimento ou 'salvar' para guardar no Drive."
            result = {
                "reply": reply,
                "delay_ms": calculate_delay_ms(reply),
                "presence": "paused",
                "metadata": {
                    "agent_id": "document-handler",
                    "response_identity": "Jennifer",
                    "waiting_confirmation": "attachment_mode",
                },
            }
            path = [{"step": 1, "phase": "pending_action", "action": "attachment_mode"}]
            return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)
        await consume_pending_action(phone, "attachment_mode")
        pending_payload = pending_action.get("payload", {}).get("attachment_payload", {})
        pending_payload["phone"] = phone
        pending_payload["instance"] = pending_payload.get("instance") or instance
        intent = {"is_attachment_save": is_save, "is_attachment_file": is_file, "is_attachment": True}
        attachment_result = await _handle_attachment(pending_payload, intent, sender_name)
        if attachment_result is not None:
            path = [{"step": 1, "phase": "pending_action", "action": "attachment_mode"},
                    {"step": 2, "phase": "attachment_handler",
                     "status": attachment_result.get("metadata", {}).get("attachment", "unknown")}]
            return await _finalize_orchestration(payload, masked_text, sender_name, attachment_result, path, cache_key)

    if confirmation is not None:
        if pending_action and pending_action.get("action_type") == "nickname_consent":
            await consume_pending_action(phone, "nickname_consent")
            action_payload = pending_action.get("payload", {})
            name = action_payload.get("first_name") or first_name
            nickname = action_payload.get("nickname", "")
            from tools.nickname import set_consent
            consent = await set_consent(phone, name, nickname, confirmation)
            reply = (f"Combinado, {nickname}! Vou usar esse apelido daqui pra frente."
                     if confirmation else f"Tudo certo, {name}. Vou continuar usando seu primeiro nome.")
            result = {"reply": reply, "delay_ms": calculate_delay_ms(reply), "presence": calculate_presence(),
                      "metadata": {"agent_id": "agent-intimacy", "response_identity": "Jennifer",
                                   "pending_action": "nickname_consent", "accepted": confirmation,
                                   "consent_recorded": "error" not in consent}}
            path = [{"step": 1, "phase": "pending_action", "action": "nickname_consent"}]
            return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)
        if pending_action and pending_action.get("action_type") == "group_consent":
            await consume_pending_action(phone, "group_consent")
            action_payload = pending_action.get("payload", {})
            group_jid = action_payload.get("group_jid", "")
            requested_intent = action_payload.get("intent", "calendar")
            original_text = action_payload.get("original_text", "")
            if original_text:
                payload["text"] = original_text
                masked_text = mask_pii(original_text)
            try:
                from tools.group import set_member_confirmation
                await set_member_confirmation(group_jid, phone, True)
            except Exception as exc:
                logger.warning("set_member_confirmation failed: %s", exc)
            reply = "Ok, liberei o acesso."
            pipeline_map = {"calendar": "calendar", "email": "email", "drive": "doc"}
            pipe_name = pipeline_map.get(requested_intent)
            if pipe_name == "calendar":
                from pipelines.calendar_pipeline import run as run_cal
                result = await run_cal(payload)
            elif pipe_name == "email":
                from pipelines.email_pipeline import run as run_email
                result = await run_email(payload)
            elif pipe_name == "doc":
                from pipelines.doc_pipeline import run as run_doc
                result = await run_doc(payload)
            else:
                result = {"reply": reply, "delay_ms": 0, "presence": "composing",
                          "metadata": {"agent_id": "group_consent", "pending_action": "group_consent"}}
            return await _finalize_orchestration(payload, masked_text, sender_name, result,
                                                  [{"step": 1, "phase": "pending_action", "action": "group_consent"}],
                                                  cache_key)

    # ========================
    # COMMANDS (before pipelines)
    # ========================
    cmd = detect_command(masked_text)
    if cmd:
        logger.info(f"Proactive command detected from {phone}: {cmd}")
        cmd_result = await apply_command(phone, cmd)
        log_action(actor="user", action="PROACTIVE_COMMAND", target=phone,
                    details={"command": cmd, "result": cmd_result})
        result = {"reply": cmd_result.get("message", "Comando aplicado."), "delay_ms": 0, "presence": "paused",
                  "metadata": {"agent_id": "command-handler", "command": cmd, "applied": True}}
        return await _finalize_orchestration(payload, masked_text, sender_name, result,
                                              [{"step": 1, "phase": "command", "command": cmd}], cache_key)

    # ========================
    # ATTACHMENT (has_document flag)
    # ========================
    if extra.get("has_document"):
        intent = {"is_attachment_save": False, "is_attachment_file": False, "is_attachment": True}
        attachment_result = await _handle_attachment(payload, intent, sender_name)
        if attachment_result is not None:
            path = [{"step": 1, "phase": "attachment_handler",
                     "status": attachment_result.get("metadata", {}).get("attachment", "unknown")}]
            return await _finalize_orchestration(payload, masked_text, sender_name, attachment_result, path, cache_key)

    # ========================
    # TIER 1: Security / Blocking (first-match, zero LLM)
    # ========================
    if _detect_intimacy(masked_text):
        return await _handle_intimacy(payload, masked_text, sender_name, cache_key, first_name, phone)

    if _detect_runtime_status(masked_text):
        result = await _handle_runtime_status(payload, instance, phone)
        path = [{"step": 1, "phase": "runtime_status"}]
        return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)

    if _detect_correction(masked_text):
        return await _handle_correction(payload, masked_text, sender_name, cache_key)

    if _detect_morality(masked_text):
        return await _handle_morality(payload, masked_text, sender_name, cache_key, instance, phone)

    # ========================
    # TIER 1.5: Auto-Discovery (DeepAgents hardcoded)
    # ========================
    # GUARDRAIL (18/08/2026): auto-discovery via deepagent_layer.get_deep_agent.
    # _detect_dynamic_toolkit retorna o slug (linkedin, youtube, etc.);
    # construimos um dict-wrapping com id do manager e delegamos para
    # _execute_agent, que sabe lidar com manager_id via deepagent_path.
    # Antes usava dynamic_factory.get_or_create() + dict(agent) que
    # quebrava com TypeError: 'CompiledStateGraph' object is not iterable.
    try:
        toolkit_slug = _detect_dynamic_toolkit(text)
        if toolkit_slug:
            manager_id = f"manager-{toolkit_slug}"
            from deepagent_layer.agents import get_deep_agent, MANAGER_PROMPTS
            if manager_id in MANAGER_PROMPTS and get_deep_agent(manager_id) is not None:
                # GUARDRAIL (21/08/2026): log estruturado do dispatch para
                # diagnostico. Antes o roteamento tier15 nao emitia nenhum
                # log, tornando impossivel saber se "qual meu perfil no
                # linkedin" chegou ao manager-linkedin ou caiu no tier LLM.
                logger.info(
                    "tier15_dispatch toolkit=%s manager=%s phone_suffix=%s",
                    toolkit_slug,
                    manager_id,
                    phone[-4:] if phone else "----",
                )
                result = await _execute_agent(
                    {"id": manager_id, "system_prompt": ""},
                    masked_text, payload, payload.get("extra", {}),
                )
                path = [
                    {"step": 1, "phase": "tier15_dispatch", "toolkit": toolkit_slug, "manager": manager_id},
                ]
                return await _finalize_orchestration(
                    payload, masked_text, sender_name, result, path, cache_key,
                )
            else:
                # Manager nao disponivel em MANAGER_PROMPTS (slug nao reconhecido).
                meta = api_registry.get_meta(toolkit_slug)
                if meta is None:
                    blocked_msg = (
                        f"⚠️ O toolkit '{toolkit_slug}' nao esta disponivel. "
                        f"Pode ser que ele nao esteja na allowlist "
                        f"(tools/api_registry.py::ALLOWED_TOOLKITS)."
                    )
                else:
                    blocked_msg = (
                        f"⚠️ O toolkit '{toolkit_slug}' ainda nao foi configurado pelo admin. "
                        f"O modulo tools/{toolkit_slug}_composio.py precisa existir."
                    )
                early_result = {
                    "reply": blocked_msg,
                    "delay_ms": 0,
                    "presence": "composing",
                    "metadata": {
                        "agent_id": "tier15-not-found",
                        "toolkit": toolkit_slug,
                        "blocked_reason": "not_allowed_or_module_missing",
                    },
                }
                path = [
                    {"step": 1, "phase": "tier15_blocked", "toolkit": toolkit_slug},
                ]
                return await _finalize_orchestration(
                    payload, masked_text, sender_name, early_result, path, cache_key,
                )
    except Exception as exc:
        logger.warning("tier15_dispatch_handler_failed: %s", exc)

    # ========================
    # TIER 1.7: Deterministic Pipeline Detection (zero LLM)
    # ========================
    # GUARDRAIL (17/08/2026): detectores deterministicos (calendar/email/drive)
    # rodam ANTES do classificador LLM. Resolve o bug em que pedidos claros
    # de email eram classificados como "conversa" pelo LLM e nunca chegavam
    # ao email_pipeline (ex: "leia meu email da XP sobre o processo seletivo").
    # GUARDRAIL (17/08/2026): detectores deterministicos (calendar/email/drive)
    # rodam ANTES do classificador LLM. Resolve o bug em que pedidos claros
    # de email eram classificados como "conversa" pelo LLM e nunca chegavam
    # ao email_pipeline (ex: "leia meu email da XP sobre o processo seletivo").
    # Fix E1 (18/08/2026): detectores usam o texto ORIGINAL (pre-mask) —
    # [MASK_EMAIL] anteriormente invertia o roteamento de pedidos de
    # calendario com email de participante (ex: "marque um compromisso com
    # o Maycon... invite mayconpxavier@gmail.com").
    from pipelines.calendar_pipeline import detect as cal_detect, run as cal_run
    from pipelines.email_pipeline import detect as eml_detect, run as eml_run
    from pipelines.doc_pipeline import detect_drive_attachment, run as doc_run
    from pipelines.jennifer_pipeline import run as jen_run

    if cal_detect(text):
        result = await cal_run(payload)
        path = [{"step": 1, "phase": "deterministic_routing", "detector": "calendar"}]
        return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)
    if eml_detect(text):
        result = await eml_run(payload)
        path = [{"step": 1, "phase": "deterministic_routing", "detector": "email"}]
        return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)
    if detect_drive_attachment(text):
        result = await doc_run(payload)
        path = [{"step": 1, "phase": "deterministic_routing", "detector": "drive"}]
        return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)

    # ========================
    # TIER 2: LLM Classifier (Flash, 1 call, ~200ms) — fallback
    # ========================
    intent_class = await _classify_intent_llm(masked_text)

    if intent_class in ("juridicas", "editais", "academica", "anotacoes"):
        payload["intent_class"] = intent_class
        result = await doc_run(payload)
    elif intent_class == "ferramentas":
        # Detectores deterministicos ja rodaram acima (TIER 1.7). Se chegou
        # aqui, e intencao de ferramenta nao capturada por keywords.
        result = await jen_run(payload)
    else:
        await _setup_nickname_consent(payload, first_name, phone)
        result = await jen_run(payload)

    path = [{"step": 1, "phase": "pipeline_routing", "intent_class": intent_class}]
    return await _finalize_orchestration(payload, masked_text, sender_name, result, path, cache_key)


_LEGACY_TOOL_CALL_RE = re.compile(r"<\s*tool_call\s*>.*?</\s*tool_call\s*>", re.DOTALL)
_LEGACY_TOOL_CALL_SELF_RE = re.compile(r"<\s*tool_call\s*/?\s*>")
_LEGACY_INVOKE_RE = re.compile(r"<\s*invoke\b[^>]*>.*?</\s*invoke\s*>", re.DOTALL)
_LEGACY_INVOKE_SELF_RE = re.compile(r"<\s*invoke\b[^>]*/?>")


def _strip_provider_artifacts(text: str) -> str:
    """Safety net: remove generic tool_call/invoke tags that may leak into `content`.

    `chat_with_tools` strips these when it parses inline tool calls, but if
    the parser is bypassed (e.g. provider returns clean `tool_calls`), a few
    fragments can still survive. We collapse them here as defense in depth.

    NOTE: legacy MiniMax-specific tags (``[<minimax>[``, ``]<minimax>]``) were
    removed 15/08/2026 along with the MiniMax provider (Fase N consolidated
    DeepSeek V4 Flash as the sole LLM).
    """
    cleaned = _LEGACY_TOOL_CALL_RE.sub("", text)
    cleaned = _LEGACY_TOOL_CALL_SELF_RE.sub("", cleaned)
    cleaned = _LEGACY_INVOKE_RE.sub("", cleaned)
    cleaned = _LEGACY_INVOKE_SELF_RE.sub("", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _normalize_response_identity(text: str) -> str:
    normalized = _strip_provider_artifacts(str(text or ""))
    replacements = [
        (r"(?i)\bsou\s+(?:o|a)\s+(?:web|calendar|drive|email)\s+manager\b", "sou a Jennifer"),
        (r"(?i)\baqui\s+e\s+(?:o|a)\s+(?:web|calendar|drive|email)\s+manager\b", "aqui e a Jennifer"),
        (r"(?i)\b(?:sua|seu)\s+(?:web|calendar|drive|email)\s+manager\b", "Jennifer"),
        (r"(?i)\b(?:web|calendar|drive|email)\s+manager\b", "Jennifer"),
        (r"(?i)\bmanager-(?:web|calendar|drive|email)\b", "Jennifer"),
        (r"(?i)\bagent-(?:intimacy|learning|morality|rag)\b", "Jennifer"),
    ]
    for pattern, replacement in replacements:
        normalized = re.sub(pattern, replacement, normalized)
    return normalized


def _bind_tool_args(tool_name: str, tool_args: Dict[str, Any], phone: str, instance: str = "", remote_jid: str = "") -> Dict[str, Any]:
    effective_args = dict(tool_args)
    if is_user_scoped_tool(tool_name):
        fn = get_tool(tool_name)
        if _fn_accepts_kwarg(fn, "phone"):
            effective_args["phone"] = phone
        if _fn_accepts_kwarg(fn, "instance"):
            effective_args["instance"] = instance

    if tool_name.startswith("group."):
        current_gjid = effective_args.get("group_jid") or ""
        if ("@g.us" not in str(current_gjid)) and ("@g.us" in str(remote_jid)):
            effective_args["group_jid"] = remote_jid
    return effective_args


def _fn_accepts_kwarg(fn, name: str) -> bool:
    """True se ``fn`` aceita o kwarg ``name`` (parametro explicito ou **kwargs).

    Desacopla o binding de tools da assinatura de cada implementacao: tools
    user-scoped com assinatura estrita (ex: people.search, tasks.list) NAO
    recebem ``instance``, evitando TypeError. Fallback conservador: se nao
    for possivel inspecionar a assinatura, injeta (comportamento antigo).
    """
    if fn is None:
        return True
    import inspect

    try:
        params = inspect.signature(fn).parameters
    except (TypeError, ValueError):
        return True
    if name in params:
        return True
    return any(p.kind == inspect.Parameter.VAR_KEYWORD for p in params.values())


def _is_ai_message(message: Any) -> bool:
    """Return True if the message is an AI assistant message.

    Compatible with LangChain 1.x ``AIMessage`` (BaseMessage subclass)
    and the legacy dict format used by early DeepAgents.
    """
    if isinstance(message, dict):
        return message.get("role") == "assistant"
    cls_name = type(message).__name__
    if cls_name == "AIMessage":
        return True
    msg_type = getattr(message, "type", "")
    return msg_type in {"ai", "AIMessage"}


def _is_tool_message(message: Any) -> bool:
    """Return True if the message is a tool result message."""
    if isinstance(message, dict):
        return message.get("role") == "tool"
    cls_name = type(message).__name__
    if cls_name == "ToolMessage":
        return True
    msg_type = getattr(message, "type", "")
    return msg_type in {"tool", "ToolMessage"}


def _tool_message_name(message: Any) -> str:
    """Extract the tool name associated with a tool result message.

    Returns empty string when the message has no ``name`` so the caller
    can pair it with the earliest pending AI tool_call. Returns the
    ``tool_call_id`` only as a last-resort fallback (rarely needed in
    practice but keeps the tool_results list ordered when LangChain
    strips the name).
    """
    if isinstance(message, dict):
        return str(message.get("name") or message.get("tool") or "")
    name = getattr(message, "name", None)
    if name:
        return str(name)
    return ""


def _tool_message_payload(message: Any) -> Any:
    """Extract the actual result data from a tool message.

    LangChain ``ToolMessage.content`` can be a string, a list of
    blocks, or arbitrary JSON. Pure-dict messages (no LangChain
    objects) carry ``content`` directly. We try to normalise to a
    Python dict/list when the content is a JSON string.
    """
    if isinstance(message, dict):
        content = message.get("content")
        if content is None:
            content = message.get("data")
    else:
        content = getattr(message, "content", None)
    if content is None:
        return None
    if isinstance(content, (dict, list)):
        return content
    if isinstance(content, str):
        try:
            import json as _json
            return _json.loads(content)
        except Exception:
            return content
    return content


def _extract_deepagent_tool_results(messages: List[Any]) -> List[Dict[str, Any]]:
    """Walk the DeepAgent message log and pair tool calls with their
    results. Returns a list of ``{"tool": str, "result": Any}`` items
    in chronological order, ready for ``_detect_tabular_payload``.

    Robust to LangChain 1.x ``AIMessage.tool_calls`` (list of dicts)
    and to messages without explicit tool_calls (fallback to the
    ``ToolMessage.name`` field).
    """
    results: List[Dict[str, Any]] = []
    pending_calls: List[str] = []

    for m in messages or []:
        if _is_ai_message(m):
            tool_calls = getattr(m, "tool_calls", None)
            if isinstance(tool_calls, list):
                for call in tool_calls:
                    if isinstance(call, dict):
                        name = call.get("name") or call.get("tool") or ""
                    else:
                        name = getattr(call, "name", "") or ""
                    if name:
                        pending_calls.append(str(name))
        elif _is_tool_message(m):
            tool_name = _tool_message_name(m)
            if not tool_name and pending_calls:
                tool_name = pending_calls.pop(0)
            payload = _tool_message_payload(m)
            if isinstance(payload, (dict, list)):
                results.append({"tool": tool_name or "unknown_tool",
                                 "result": payload})
            else:
                results.append({"tool": tool_name or "unknown_tool",
                                 "result": {"raw": str(payload) if payload is not None else ""}})
    return results[-10:]


def _extract_message_content(message: Any) -> str:
    """Extract the text content from an AI message in any supported format.

    LangChain 1.x ``AIMessage`` may carry content as a string OR as a list
    of blocks (e.g. ``[{"type": "text", "text": "..."}]``). We normalise to
    a plain string.
    """
    if isinstance(message, dict):
        content = message.get("content", "")
    else:
        content = getattr(message, "content", "")
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts = []
        for block in content:
            if isinstance(block, dict):
                text = block.get("text", "")
                if text:
                    parts.append(str(text))
            elif isinstance(block, str):
                parts.append(block)
        return " ".join(parts)
    return str(content or "")


async def _execute_deep_agent(
    agent: Dict[str, Any],
    text: str,
    payload: Dict[str, Any],
    extra: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    """Execute the agent using DeepAgents harness.

    Returns ``None`` if DeepAgents is not available for this agent id
    (e.g. ``manager-calendar`` is supported, but a custom agent id is not).
    Returns the standard reply dict on success, or a fallback error dict
    on failure.
    """
    agent_id = agent.get("id", "unknown-agent")
    try:
        from deepagent_layer import get_deep_agent
    except Exception as exc:
        logger.warning("deepagent_layer unavailable: %s", exc)
        return None

    with current_tracker().stage("deepagent_build", agent_id=agent_id):
        deep_agent = get_deep_agent(agent_id)
    if deep_agent is None:
        return None

    from core.agent_status import record_agent_failure, record_agent_success, start_agent_execution
    execution_started = start_agent_execution(agent_id)

    phone = payload.get("phone", "")
    first_name = payload.get("first_name", "")
    brt = timezone(timedelta(hours=-3))
    hoje = datetime.now(brt)

    user_prompt = (
        f"User: {payload.get('sender_name', 'user')} (tel: {phone}"
        + (f", primeiro nome: {first_name}" if first_name else "")
        + ")\n"
        f"Mensagem: {text}\n\n"
        f"DATA ATUAL: {hoje.strftime('%Y-%m-%d')} (horario de Brasilia, BRT, UTC-3). "
        f"Hora atual: {hoje.strftime('%H:%M')}. "
        f"Seja proativa: se a pergunta for ambigua, pergunte antes de agir. "
        f"Ofereca opcoes em vez de presumir. Lembre do contexto da conversa. "
        f"Responda em portugues brasileiro, tom caloroso, natural."
    )

    config: Dict[str, Any] = {
        "configurable": {
            "thread_id": phone or "default",
        }
    }
    if phone:
        config["configurable"]["phone"] = phone
    instance = payload.get("instance", "")
    if instance:
        config["configurable"]["instance"] = instance
        config["configurable"]["_instance"] = instance
    from core.runtime_context import set_instance, set_phone
    set_instance(instance)
    set_phone(phone)

    try:
        with current_tracker().stage("deepagent_ainvoke", agent_id=agent_id):
            result = await asyncio.wait_for(
                deep_agent.ainvoke(
                    {"messages": [{"role": "user", "content": user_prompt}]},
                    config=config,
                ),
                timeout=120,
            )
    except asyncio.TimeoutError:
        record_agent_failure(agent_id, execution_started, "deepagent_timeout")
        return _error_response(504, "deepagent_timeout", "Jennifer demorou demais. Tenta de novo?")
    except Exception as exc:
        record_agent_failure(agent_id, execution_started, f"deepagent_error:{type(exc).__name__}")
        logger.exception("deepagent_failed manager=%s", agent_id)
        return None

    messages = (result or {}).get("messages", [])
    reply_text = ""
    for m in reversed(messages):
        if _is_ai_message(m):
            content = _extract_message_content(m)
            reply_text = str(content or "").strip()
            break

    if not reply_text:
        record_agent_failure(agent_id, execution_started, "deepagent_empty")
        return _error_response(500, "deepagent_empty", "Nao consegui gerar uma resposta.")

    reply_text = re.sub(r'\s*<think>.*?</think>\s*', '', reply_text, flags=re.DOTALL).strip()
    reply_text = _normalize_response_identity(reply_text)
    delay_ms = calculate_delay_ms(reply_text)
    presence = calculate_presence()

    captured_tool_results = _extract_deepagent_tool_results(messages)

    from core import metrics
    model_used = "deepseek-v4-flash"
    provider = "deepseek"
    metrics.record_provider_latency(provider, True, execution_started)
    record_agent_success(agent_id, execution_started, model_used, provider)

    return {
        "reply": reply_text,
        "delay_ms": delay_ms,
        "presence": presence,
        "metadata": {
            "agent_id": agent_id,
            "executed_agent_id": agent_id,
            "response_identity": "Jennifer",
            "model_used": model_used,
            "provider": provider,
            "tool_rounds": len([m for m in messages if _is_tool_message(m)]),
            "tool_calls": [],
            "tool_results": captured_tool_results,
            "has_audio": extra.get("has_audio", False),
            "runtime": "deepagents",
        },
    }


async def _verify_calendar_event(phone: str, result: dict, tool_args: dict) -> dict:
    """Anti-alucinacao: confirma que um evento criado realmente existe.

    Apos calendar.create_event retornar sucesso, consulta a agenda nas
    proximas 48h e verifica se o evento com o summary esperado apareceu.
    Se nao aparecer, devolve erro explicito para o LLM nao declarar
    'criei/agendei' sem base real.
    """
    try:
        from tools.google_calendar import list_events
        from core.timezone import now_brt
        from datetime import timedelta

        summary = str(tool_args.get("summary") or result.get("summary") or "")
        start = str(tool_args.get("start") or result.get("start") or "")
        now = now_brt()
        listing = await list_events(
            phone,
            time_min=now.isoformat(),
            time_max=(now + timedelta(hours=48)).isoformat(),
            max_results=50,
        )
        events = listing.get("events", []) if isinstance(listing, dict) else []
        if not summary:
            return result
        match = any(str(e.get("summary", "")) == summary for e in events)
        if not match:
            logger.warning(
                "create_event_false_positive phone=%s summary=%s start=%s events_seen=%d",
                phone, summary, start, len(events),
            )
            return {
                **result,
                "error": (
                    f"O evento '{summary}' NAO foi encontrado na agenda apos a criacao. "
                    "NAO diga que criou. Informe o usuario que a criacao falhou."
                ),
            }
        return result
    except Exception as exc:  # noqa: BLE001
        logger.debug("verify_calendar_event_skipped phone=%s exc=%s", phone, exc)
        return result


def _onboarding_url(phone: str) -> str:
    """Gera URL publica de onboarding/conexão com Magic Link assinado para o user."""
    try:
        from core.magic_link import build_magic_link_url

        raw_url = build_magic_link_url(phone)
        try:
            from core.link_shortener import shorten_urls_in_text

            return shorten_urls_in_text(raw_url)
        except Exception:
            return raw_url
    except Exception:
        digits = "".join(c for c in str(phone) if c.isdigit())
        return f"/portal/?phone={digits}"


def _resolve_agent_tools(agent: Dict[str, Any]) -> List[str]:
    """Resolve as tools disponiveis para um agente (fix 12/08/2026).

    Dinamico: se o agente NAO define uma lista explicita de tools, expoe
    TODAS as tools registradas no tool_registry (inclui composio.*,
    youtube.*, locomotion.*, weather.*, etc). Assim, conectar um app no
    Composio libera a tool automaticamente, sem backfill/seed manual.
    - tools: None/ausente -> todas as tools do registry
    - tools: [...] -> apenas a lista explicita
    - tools: [] -> nenhuma tool (bloqueio total)
    """
    explicit = agent.get("tools")
    if explicit is None:
        from tool_registry import list_llm_tool_ids

        return list_llm_tool_ids()
    return [str(t) for t in explicit]


async def _execute_agent(
    agent: Dict[str, Any],
    text: str,
    payload: Dict[str, Any],
    extra: Dict[str, Any],
) -> Dict[str, Any]:
    """Execute a specific agent with tool calling loop.

    Tries the DeepAgents harness first; falls back to the legacy
    LLMProvider path if DeepAgents is unavailable for this agent id.
    """
    agent_id = agent.get("id", "unknown-agent")
    from core.agent_status import record_agent_failure, record_agent_success, start_agent_execution

    execution_started = start_agent_execution(agent_id)

    # ── Skills section (guarded: se falhar, agente segue sem skills) ──
    skills_section = ""
    try:
        skills_section = _build_skills_section(agent.get("skills", []))
    except Exception as exc:
        logger.warning("skills_section_failed agent_id=%s exc=%s", agent_id, exc)

    system_prompt = agent.get("system_prompt", "") + skills_section
    if agent.get("role") != "orchestrator":
        system_prompt += (
            "\n\n[IDENTIDADE EXTERNA OBRIGATORIA]\n"
            "Voce e um componente interno da Jennifer. Nunca revele nome, ID, role ou arquitetura interna. "
            "Responda ao usuario sempre na voz da Jennifer e nunca se apresente como Manager ou Specialist."
        )

    phone = payload.get("phone", "")
    user_context_parts: List[str] = []

    if phone:
        try:
            from tools.correction import summarize_past_corrections
            corr = await summarize_past_corrections(phone, limit=3)
            if corr.get("has_corrections"):
                items = corr["corrections"]
                corr_str = "\n".join(
                    f"- Correcao anterior ({c['target']}): '{c['user_quote'][:100]}' → '{c['after'][:100]}'"
                    for c in items
                )
                user_context_parts.append(f"[APRENDIZADOS DO USUARIO]\n{corr_str}\nRespeite essas preferencias ao responder.")
        except Exception:
            pass

    history = ""
    try:
        history = await _get_context_for_prompt(phone, limit=10) or ""
    except Exception as exc:
        logger.warning("history_fetch_failed agent_id=%s exc=%s", agent_id, exc)

    facts = ""
    try:
        if phone:
            from tools.memory import search_facts
            fact_result = await search_facts(query="", phone=phone, limit=30)
            fact_items = fact_result.get("results", []) if isinstance(fact_result, dict) else []
            if fact_items:
                facts = "\n".join(
                    f"- {f.get('key', '')}: {f.get('value', '')}"
                    for f in fact_items
                )
    except Exception as exc:
        logger.warning("memory_facts_failed agent_id=%s exc=%s", agent_id, exc)

    group_ctx = ""
    mention_ctx = ""
    current_group_ctx = ""
    try:
        if phone:
            group_ctx = await _user_groups_context(phone)
        if extra.get("is_group"):
            mention_ctx = await _resolve_group_mentions(payload)
            remote_jid = str(payload.get("remote_jid") or extra.get("remote_jid") or "")
            group_name = extra.get("group_name") or ""
            if not group_name and remote_jid:
                group_name = _get_group_name_by_jid(phone, remote_jid)
            g_name_str = group_name or "Grupo Atual"
            current_group_ctx = (
                f"[GRUPO ATUAL DO WHATSAPP]\n"
                f"Voce esta respondendo DENTRO do grupo do WhatsApp: '{g_name_str}' (JID: {remote_jid}).\n"
                f"- PAPEL: Voce atua neste grupo ESTRITAMENTE como SECRETARIA EXECUTIVA DA COHERENCE (agenda, reunioes, e-mails, documentos, atas, tarefas e rotinas da equipe).\n"
                f"- GUARDRAIL DE ESCOPO & FINOPS: Voce NAO e um ChatGPT para tirar duvidas de conhecimentos gerais, resolver questoes escolares, licoes de casa, matematica/trigonometria, curiosidades aleatorias ou enciclopedia. "
                f"Se membros do grupo fizerem perguntas fora do escopo corporativo (ex: 'quem teorizou a evolucao', 'o que e bitcoin', 'calcule raiz quadrada/seno/cosseno', 'piadas'), "
                f"recuse educadamente em 1 a 2 frases curtas: 'Como secretária executiva da equipe, meu foco por aqui é ajudar com agenda, e-mails, atas, documentos e tarefas de trabalho. Para dúvidas de conhecimentos gerais ou estudos, recomendo ferramentas de busca! Em que posso ajudar no trabalho hoje?'.\n"
                f"- NUNCA gere redacoes longas ou tratados cientificos para perguntas fora de escopo.\n"
                f"- Quando o usuario perguntar 'qual grupo voce esta', 'onde estou mandando mensagem' ou pedir acoes/cumprimentos no grupo atual, "
                f"responda que voce esta no grupo '{g_name_str}'.\n"
                f"- NUNCA confunda o ID de um usuario/bot (ex: 75793925419076) com o ID do grupo.\n"
                f"- NUNCA responda com relatorios tecnicos de banco de dados (ex: 'sincronizacao foi concluida para X grupos'). Responda em tom amigavel, humano e natural."
            )

            if remote_jid and phone:
                sender_name = str(payload.get("sender_name") or "")
                if sender_name and sender_name != "user":
                    from tools.group import enrich_member_name
                    asyncio.create_task(enrich_member_name(remote_jid, phone, sender_name))
    except Exception as exc:  # noqa: BLE001
        logger.warning("user_groups_context_failed agent_id=%s exc=%s", agent_id, exc)

    mem_rag = ""
    try:
        mem_rag = await _search_memory(phone, text, limit=5) or ""
    except Exception as exc:  # noqa: BLE001
        logger.warning("memory_rag_failed agent_id=%s exc=%s", agent_id, exc)

    recent = [i for i in _interaction_history[-4:] if i.get("phone") == phone]
    if current_group_ctx:
        user_context_parts.append(current_group_ctx)
    if facts:
        user_context_parts.append(f"[FATOS DO USUARIO - NAO perguntar novamente]\n{facts}")
    if group_ctx:
        user_context_parts.append(group_ctx)
    if mention_ctx:
        user_context_parts.append(mention_ctx)
    if mem_rag:
        user_context_parts.append(f"[MEMORIA RAG - CONVERSAS RELEVANTES]\n{mem_rag}")
    if recent:
        user_context_parts.append("\n".join(f"- User: {r['text_preview'][:60]}\n- Jennifer: {r['reply_preview'][:60]}"
                                            for r in recent[-2:]))
    if history:
        user_context_parts.append(f"[HISTORICO RECENTE]\n{history}")

    try:
        deep_result = await _execute_deep_agent(agent, text, payload, extra)
        if deep_result is not None:
            return deep_result
    except Exception as exc:
        logger.warning("deepagent_attempt_failed agent_id=%s exc=%s", agent_id, type(exc).__name__)

    brt = timezone(timedelta(hours=-3))
    hoje = datetime.now(brt)
    temporal_context = (
        f"[DATA ATUAL: {hoje.strftime('%Y-%m-%d')} (horario de Brasilia, BRT, UTC-3). "
        f"Hora atual: {hoje.strftime('%H:%M')}. "
        "Use esta data para todas as consultas de calendario e referencias temporais. "
        "IDIOMA: SEMPRE responda em portugues brasileiro (pt-BR). NAO use ingles. "
        "NAO inclua tags XML como <think> nas suas respostas.]"
    )

    first_name = payload.get("first_name", "")
    static_user_prefix = (
        f"User: {payload.get('sender_name', 'user')} (tel: {payload.get('phone', '')}"
        + (f", primeiro nome: {first_name}" if first_name else "")
        + ")\n"
    )

    dynamic_context_str = "\n\n".join(user_context_parts)
    if dynamic_context_str:
        user_prompt = f"{temporal_context}\n\n[CONTEXTO DA CONVERSA]\n{dynamic_context_str}\n\n{static_user_prefix}Mensagem: {text}"
    else:
        user_prompt = f"{temporal_context}\n\n{static_user_prefix}Mensagem: {text}"

    available_tools = _resolve_agent_tools(agent)

    llm = LLMProvider()
    if not llm.is_available():
        record_agent_failure(agent_id, execution_started, "llm_unavailable")
        return _error_response(503, "llm_unavailable", "Nenhum provedor LLM configurado")

    thinking = agent.get("thinking", "disabled") == "enabled"
    fast_model = agent.get("model", "deepseek-v4-flash")

    tool_schemas = []
    for tid in available_tools:
        schema = get_tool_schema(tid)
        if schema:
            tool_schemas.append({"type": "function", "function": schema})

    captured_tool_results: List[Dict[str, Any]] = []

    async def tool_executor(tool_name: str, tool_args: dict) -> str:
        _tool_started = time.monotonic()
        tool_fn = get_tool(tool_name)
        if not tool_fn:
            logger.warning("tool_unknown tool=%s", tool_name)
            return json.dumps({"error": f"Tool '{tool_name}' not found"})
        effective_args = _bind_tool_args(tool_name, tool_args, phone, payload.get("instance", ""))
        logger.info("tool_invoking tool=%s args=%s", tool_name, list(effective_args.keys()))

        try:
            from pipelines._ack import send_instant_tool_ack
            asyncio.create_task(
                send_instant_tool_ack(
                    tool_name=tool_name,
                    phone=phone,
                    instance=str(payload.get("instance", "") or "Jennifer"),
                    extra=extra,
                )
            )
        except Exception:
            pass

        try:
            coro = tool_fn(**effective_args)
            if asyncio.iscoroutine(coro) or asyncio.iscoroutinefunction(tool_fn):
                result = await asyncio.wait_for(coro, timeout=30)
            else:
                result = coro
            # Anti-alucinacao (fix 12/08/2026): apos calendar.create_event
            # "bem-sucedido", verificar se o evento realmente apareceu na
            # agenda. Evita que o LLM diga "criei/agendei" sem confirmacao.
            if tool_name == "calendar.create_event" and isinstance(result, dict) and not result.get("error"):
                result = await _verify_calendar_event(phone, result, tool_args)
            if isinstance(result, dict):
                captured_tool_results.append({"tool": tool_name, "result": result})
                if len(captured_tool_results) > 10:
                    captured_tool_results.pop(0)
            truncated = mask_pii(json.dumps(result, ensure_ascii=False, default=str))
            if len(truncated) > 2000:
                truncated = truncated[:2000] + "...(truncated)"
            logger.info(
                "tool_invocation_result tool=%s status=%s duration_ms=%d",
                tool_name,
                "ok" if (isinstance(result, dict) and not result.get("error")) else "ok",
                round((time.monotonic() - _tool_started) * 1000, 2),
            )
            logger.info("tool_result tool=%s length=%d", tool_name, len(truncated))
            return truncated
        except asyncio.TimeoutError:
            logger.error("tool_timeout tool=%s timeout=30s", tool_name)
            return json.dumps({"error": f"Tool '{tool_name}' timed out after 30s"})
        except Exception as e:
            logger.exception("tool_error tool=%s", tool_name)
            err_msg = f"{type(e).__name__}: {str(e)[:200]}"
            # Onboarding: quando a tool precisa de OAuth Google ou Composio, devolve
            # mensagem com link limpo encurtado para o user autorizar de forma autonoma (sem imagens).
            err_lower = err_msg.lower()
            if (
                "oauth_required" in err_lower
                or "user_google_oauth_required" in err_lower
                or "composio_oauth_required" in err_lower
                or "unregistered callers" in err_lower
                or "no active connection found" in err_lower
                or "needs_connection" in err_lower
                or "folder_permission_required" in err_lower
                or "scope_missing" in err_lower
            ):
                return json.dumps({
                    "error": (
                        "acao_requer_autorizacao",
                        f"O usuario precisa conectar ou autorizar seu servico. "
                        f"Envie uma mensagem curta e amigavel (sem imagens) com o link: 'Para eu acessar este servico por voce, conecte suas contas aqui: {_onboarding_url(phone)}'",
                    ),
                })
            return json.dumps({"error": err_msg})

    try:
        if tool_schemas:
            with current_tracker().stage("llm_chat_with_tools", model=fast_model):
                result = await llm.chat_with_tools(
                    system_prompt=system_prompt,
                    user_prompt=user_prompt,
                    tools=tool_schemas,
                    tool_executor=tool_executor,
                    model=fast_model,
                    temperature=0.7,
                    max_tokens=LLM_MAX_TOKENS_MANAGER,
                    thinking_disabled=not thinking,
                    max_tool_rounds=5,
                )
        else:
            with current_tracker().stage("llm_chat", model=fast_model):
                result = await llm.chat(
                    system_prompt=system_prompt,
                    user_prompt=user_prompt,
                    model=fast_model,
                    temperature=0.7,
                    max_tokens=LLM_MAX_TOKENS_DEFAULT,
                    thinking_disabled=not thinking,
                )

        usage = result.get("usage") or {}
        if usage:
            current_tracker().add_costs(
                deepseek_input_tokens=usage.get("prompt_tokens", 0),
                deepseek_output_tokens=usage.get("completion_tokens", 0),
                deepseek_cache_hit_tokens=usage.get("cache_hit_tokens", 0),
            )

        reply_text = result["content"]
        reply_text = re.sub(r'\s*<think>.*?</think>\s*', '', reply_text, flags=re.DOTALL).strip()
        reply_text = _normalize_response_identity(reply_text)
        delay_ms = calculate_delay_ms(reply_text)
        presence = calculate_presence()

        tool_calls_made = _extract_tool_calls(reply_text, available_tools)
        model_used = result.get("model_used", fast_model)
        provider = result.get("provider") or "unknown"
        if provider == "unknown":
            lowered = model_used.lower()
            if "deepseek" in lowered:
                provider = "deepseek"
            elif "gemini" in lowered:
                provider = "gemini"
            elif "groq" in lowered:
                provider = "groq"
        from core import metrics

        metrics.record_provider_latency(provider, True, execution_started)
        record_agent_success(agent_id, execution_started, model_used, provider)

        return {
            "reply": reply_text,
            "delay_ms": delay_ms,
            "presence": presence,
            "metadata": {
                "agent_id": agent_id,
                "executed_agent_id": agent_id,
                "response_identity": "Jennifer",
                "model_used": model_used,
                "provider": provider,
                "tool_rounds": result.get("tool_rounds", 0),
                "tool_calls": tool_calls_made,
                "tool_results": captured_tool_results,
                "has_audio": extra.get("has_audio", False),
            },
        }
    except LLMError as e:
        record_agent_failure(agent_id, execution_started, str(e))
        logger.error(f"LLM cascade failed for agent {agent_id}: {e}")
        return _error_response(503, "llm_unavailable", "Todos provedores LLM falharam.")
    except Exception as e:
        record_agent_failure(agent_id, execution_started, str(e))
        logger.exception("Unexpected error in _execute_agent")
        return _error_response(500, "internal_error", str(e))


def _extract_tool_calls(reply_text: str, available_tools: List[str]) -> List[Dict[str, Any]]:
    """Best-effort detection of tool calls mentioned in reply.

    Checks both resource name (calendar) and method (list_events).
    """
    tool_calls: List[Dict[str, Any]] = []
    if not reply_text or not available_tools:
        return tool_calls

    reply_lower = reply_text.lower()
    for tool_id in available_tools:
        parts = tool_id.split(".")
        resource = parts[0].replace("_", " ") if len(parts) > 0 else ""
        method = parts[-1].replace("_", " ") if len(parts) > 1 else ""

        matched = False
        match_type = None
        if resource and resource in reply_lower:
            matched = True
            match_type = "resource"
        elif method and method in reply_lower:
            matched = True
            match_type = "method"

        if matched:
            tool_calls.append({
                "tool_id": tool_id,
                "source": "text_match",
                "match_type": match_type,
            })
    return tool_calls


def _error_response(status_code: int, error: str, message: str) -> Dict[str, Any]:
    return {
        "reply": message,
        "delay_ms": 0,
        "presence": "paused",
        "metadata": {
            "error": error,
            "status_code": status_code,
            "response_identity": "Jennifer",
        },
    }


# async def _run_guard_graph(payload: Dict[str, Any], masked_text: str, intent: Dict[str, bool]) -> Dict[str, Any]:
#     """Run the guard pipeline for a single turn.
#
#     Fase A.2 (30/07/2026): substitui LangGraph StateGraph por chamada
#     direta aos nodes via ``run_guard_sync``. Mesma semantica, sem
#     overhead de CompiledStateGraph dispatch. Custo: ~-$0.10/mes
#     (menos CPU time em warm path).
#
#     Returns a normalized decision dict with ``verdict`` (``allow`` /
#     ``request_oauth`` / ``deny`` / ``noop``) and an optional ``reply``
#     the orchestrator should use when the verdict blocks the user.
#     """
#     try:
#         from agent_orchestration.access_guardian import decide_guardian
#         from agent_orchestration.graph import run_guard_sync
#     except Exception as exc:
#         logger.warning("agent_orchestration unavailable, skipping guard: %s", exc)
#         return {"verdict": "noop", "trace": [], "reason": f"graph_unavailable:{exc}"}
#
#     capability = _intent_to_capability(intent)
#     initial_state = {
#         "instance": payload.get("instance", ""),
#         "phone": payload.get("phone", ""),
#         "sender_name": payload.get("sender_name", ""),
#         "text": payload.get("text", ""),
#         "masked_text": masked_text,
#         "remote_jid": payload.get("extra", {}).get("remote_jid", ""),
#         "intent": dict(intent),
#         "capability": capability,
#     }
#
#     try:
#         final = await run_guard_sync(initial_state)
#     except Exception as exc:
#         logger.warning("guard sync execution failed: %s", exc)
#         decision = decide_guardian(
#             instance=payload.get("instance", ""),
#             phone=payload.get("phone", ""),
#             capability=capability or "noop",
#         )
#         return {"verdict": decision.verdict, "decision": decision.to_dict(), "reason": exc.__class__.__name__}
#
#     decision = (final or {}).get("guardian_decision") or {}
#     verdict = decision.get("verdict", "allow")
#     prefetch = (final or {}).get("prefetch")
#     return {
#         "verdict": verdict,
#         "decision": decision,
#         "prefetch": prefetch,
#         "trace": (final or {}).get("trace", []),
#     }


# def _intent_to_capability(intent: Dict[str, bool]) -> str:
#     if intent.get("is_rag"):
#         return "knowledge.retrieve"
#     if intent.get("is_calendar"):
#         return "calendar.list_events"
#     if intent.get("is_email"):
#         return "gmail.search_messages"
#     if intent.get("is_drive"):
#         return "drive.search_files"
#     return ""
